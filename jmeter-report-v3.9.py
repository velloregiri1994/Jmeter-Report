#!/usr/bin/env python3
"""
JMeter HTML Performance Dashboard (Tabs, Filters, Advanced Analysis) - v3.9
--------------------------------------------------------------------

Enhanced Features in v3.9:
1. Health Score calculation with letter grade
2. Root Cause Analysis with confidence levels
3. Priority-based recommendations (Critical/High/Medium/Low)
4. Capacity planning estimates
5. Pattern recognition for common issues
6. Actionable remediation steps
7. Container resource comparison (Before/During/After test)
8. Excel export with rich formatting and charts

Usage:
    # Generate HTML report only
    python jmeter_report_v3_9.py \
        --input results.csv \
        --output report.html \
        --test-name "My Load Test" \
        --environment "UAT"

    # Generate both HTML and Excel reports
    python jmeter_report_v3_9.py \
        --input results.csv \
        --output report.html \
        --excel-export report.xlsx \
        --infra-json cpu_data.json memory_data.json

    # Generate only Excel report
    python jmeter_report_v3_9.py \
        --input results.csv \
        --excel-export report.xlsx \
        --excel-only \
        --test-name "Production Load Test"
"""

import argparse
import csv
import json
import os
import math
import statistics
from collections import defaultdict, Counter
from dataclasses import dataclass, field
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple

# Add this after the existing imports
try:
    import yaml
    YAML_SUPPORT = True
except ImportError:
    YAML_SUPPORT = False
    print("Note: PyYAML not installed. YAML config files will not be supported.")
    print("Install with: pip install pyyaml")

# Excel export dependencies (optional)
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.utils import get_column_letter

    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False
    print("Warning: openpyxl not installed. Excel export will be disabled.")
    print("Install with: pip install openpyxl pandas")

import numpy as np
from scipy import stats
from scipy.signal import find_peaks, savgol_filter
import warnings
warnings.filterwarnings('ignore')

# =========================
# DEFAULT SLA CONFIGURATION
# =========================

DEFAULT_SLA_CONFIG = {
    "*": {
        "p95_ms": 2000.0,  # 95% of requests under 3 seconds
        "error_rate": 0.02  # <= 2% errors
        # Optionally add "avg_ms": 2000.0, "tps_min": 5.0 per label
    },
}


# =========================
# DATA MODELS
# =========================

@dataclass
class Sample:
    timestamp_ms: int
    elapsed_ms: float
    label: str
    response_code: str
    success: bool
    all_threads: Optional[int] = None
    response_message: str = ""  # NEW: capture error message/responseMessage


@dataclass
class LabelStats:
    label: str
    count: int = 0
    success_count: int = 0
    error_count: int = 0
    elapsed_values_ms: List[float] = field(default_factory=list)

    @property
    def error_rate(self) -> float:
        return self.error_count / self.count if self.count else 0.0

    def to_summary(self) -> Dict[str, Any]:
        if not self.elapsed_values_ms:
            return {
                "label": self.label,
                "count": self.count,
                "min": None,
                "max": None,
                "avg": None,
                "p50": None,
                "p90": None,
                "p95": None,
                "p99": None,
                "error_rate": self.error_rate,
            }
        values = sorted(self.elapsed_values_ms)

        def pct(p: float) -> float:
            return percentile(values, p)

        return {
            "label": self.label,
            "count": self.count,
            "min": min(values),
            "max": max(values),
            "avg": statistics.mean(values),
            "p50": pct(50),
            "p90": pct(90),
            "p95": pct(95),
            "p99": pct(99),
            "error_rate": self.error_rate,
        }


# =========================
# UTILS
# =========================

def percentile(sorted_values: List[float], p: float) -> float:
    """Compute percentile p (0-100) for a sorted list using linear interpolation."""
    if not sorted_values:
        return float("nan")
    k = (len(sorted_values) - 1) * (p / 100.0)
    f = math.floor(k)
    c = math.ceil(k)
    if f == c:
        return sorted_values[int(k)]
    d0 = sorted_values[f] * (c - k)
    d1 = sorted_values[c] * (k - f)
    return d0 + d1


def epoch_ms_to_sec_bucket(ts_ms: int) -> int:
    return int(ts_ms // 1000)


def format_ms(ms: Optional[float]) -> str:
    if ms is None or ms != ms:  # NaN check
        return "-"
    return f"{ms:.1f}"


def format_pct(p: Optional[float]) -> str:
    if p is None or p != p:
        return "-"
    return f"{p * 100:.2f}%"


def compute_apdex(values_ms: List[float], threshold_s: float) -> float:
    """
    APDEX: (Satisfied + Tolerating/2) / Total
    Satisfied: t <= T
    Tolerating: T < t <= 4T
    """
    if not values_ms:
        return float("nan")
    T_ms = threshold_s * 1000.0
    satisfied = sum(1 for v in values_ms if v <= T_ms)
    tolerating = sum(1 for v in values_ms if T_ms < v <= 4 * T_ms)
    total = len(values_ms)
    return (satisfied + tolerating / 2.0) / total if total else float("nan")


def format_duration(seconds: float) -> str:
    total = int(round(seconds))
    h = total // 3600
    m = (total % 3600) // 60
    s = total % 60
    return f"{h}h {m}m {s}s"


def calculate_trend(values: List[float]) -> float:
    """Calculate trend ratio (last/first) of values."""
    if not values or len(values) < 2:
        return 1.0
    first = statistics.mean(values[:max(1, len(values) // 10)])
    last = statistics.mean(values[-max(1, len(values) // 10):])
    return last / first if first > 0 else 1.0


# =========================
# CONFIGURATION FILE SUPPORT
# =========================

def load_config_file(config_path: str) -> dict:
    """Load configuration from JSON or YAML file."""
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            if config_path.endswith('.json'):
                return json.load(f)
            elif config_path.endswith(('.yaml', '.yml')):
                if not YAML_SUPPORT:
                    raise SystemExit("YAML support requires pyyaml. Install with: pip install pyyaml")
                return yaml.safe_load(f)
            else:
                # Try to auto-detect
                content = f.read()
                f.seek(0)
                if content.strip().startswith('{'):
                    return json.loads(content)
                else:
                    if YAML_SUPPORT:
                        return yaml.safe_load(f)
                    else:
                        raise SystemExit("File appears to be YAML but pyyaml is not installed")
    except Exception as e:
        raise SystemExit(f"Failed to load config file {config_path}: {e}")


def merge_config_with_args(config: dict, args: argparse.Namespace) -> argparse.Namespace:
    """Merge configuration file values with command line arguments."""
    import copy

    # Create a copy of args as a dictionary
    args_dict = vars(copy.copy(args))

    # Map config keys to argument names
    config_mapping = {
        'test': {
            'name': 'test_name',
            'environment': 'environment',
            'apdex_threshold': 'apdex_threshold'
        },
        'input': {
            'file': 'input',
            'baseline': 'baseline'
        },
        'output': {
            'html': 'output',
            'excel': 'excel_export'
        },
        'sla': {
            'file': 'sla_config_file',
            'json': 'sla_config_json'
        },
        'infrastructure': {
            'files': 'infra_json'
        },
        'options': {
            'excel_only': 'excel_only',
            'enable_container_analysis': 'enable_container_analysis'
        }
    }

    # Apply config values if not specified in command line
    for section, mapping in config_mapping.items():
        if section in config:
            for config_key, arg_key in mapping.items():
                if config_key in config[section]:
                    # Only use config value if command line didn't provide it
                    current_value = args_dict.get(arg_key)
                    if current_value is None or current_value == [] or current_value == '':
                        args_dict[arg_key] = config[section][config_key]

    # Handle SLA configuration - special case for inline config
    if 'sla' in config and 'config' in config['sla']:
        if args_dict.get('sla_config_file') is None and args_dict.get('sla_config_json') is None:
            # Save SLA config to temp file
            import tempfile
            with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as tmp:
                json.dump(config['sla']['config'], tmp, indent=2)
                args_dict['sla_config_file'] = tmp.name

    # Convert back to Namespace
    return argparse.Namespace(**args_dict)

# =========================
# PARSING
# =========================

def parse_jmeter_csv(path: str) -> List[Sample]:
    samples: List[Sample] = []
    with open(path, "r", newline="", encoding="utf-8", errors="ignore") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                ts = int(row.get("timeStamp") or row.get("timestamp") or 0)
            except ValueError:
                continue
            try:
                elapsed = float(row.get("elapsed") or 0)
            except ValueError:
                elapsed = 0.0

            label = row.get("label", "UNKNOWN")
            rc = str(row.get("responseCode", "")).strip()

            success_raw = str(row.get("success", "true")).strip().lower()
            success = success_raw in ("true", "1", "yes", "y")

            all_threads = None
            at = row.get("allThreads") or row.get("all_threads")
            if at is not None and at != "":
                try:
                    all_threads = int(at)
                except ValueError:
                    all_threads = None

            # NEW: best-effort capture of error/response message
            response_message = (
                    row.get("responseMessage")
                    or row.get("failureMessage")
                    or ""
            )

            samples.append(Sample(
                timestamp_ms=ts,
                elapsed_ms=elapsed,
                label=label,
                response_code=rc,
                success=success,
                all_threads=all_threads,
                response_message=response_message,
            ))
    return samples


# =========================
# METRIC COMPUTATION
# =========================

def extract_transaction_names(samples):
    """
    Return Transaction Controller names found in samples.
    We consider a sample to be a transaction controller parent sample if its
    response_message contains the text "Number of samples" (case-insensitive).
    This function always returns a list (may be empty).
    """
    txn_set = set()
    for s in samples:
        try:
            msg = getattr(s, "response_message", None)
            if msg and isinstance(msg, str) and "number of samples" in msg.lower():
                txn_set.add(s.label)
        except Exception:
            continue
    return sorted(txn_set)


def compute_metrics(samples: List[Sample], apdex_threshold_s: float) -> Dict[str, Any]:
    if not samples:
        return {}

    timestamps = [s.timestamp_ms for s in samples]
    min_ts = min(timestamps)
    max_ts = max(timestamps)
    duration_sec = max(1.0, (max_ts - min_ts) / 1000.0)

    elapsed_all = [s.elapsed_ms for s in samples]
    sorted_elapsed = sorted(elapsed_all)

    total = len(samples)
    errors = sum(1 for s in samples if not s.success)
    error_rate = errors / total if total else 0.0

    def pct(p: float) -> float:
        return percentile(sorted_elapsed, p)

    global_stats = {
        "total_requests": total,
        "duration_sec": duration_sec,
        "avg_ms": statistics.mean(sorted_elapsed),
        "min_ms": min(sorted_elapsed),
        "max_ms": max(sorted_elapsed),
        "p50_ms": pct(50),
        "p90_ms": pct(90),
        "p95_ms": pct(95),
        "p99_ms": pct(99),
        "error_rate": error_rate,
        "errors": errors,
        "successes": total - errors,
        "apdex": compute_apdex(sorted_elapsed, apdex_threshold_s),
    }

    # Per-label stats
    per_label: Dict[str, LabelStats] = {}
    for s in samples:
        if s.label not in per_label:
            per_label[s.label] = LabelStats(label=s.label)
        ls = per_label[s.label]
        ls.count += 1
        if s.success:
            ls.success_count += 1
        else:
            ls.error_count += 1
        ls.elapsed_values_ms.append(s.elapsed_ms)

    per_label_summaries = [ls.to_summary() for ls in per_label.values()]
    per_label_summaries.sort(key=lambda x: x["label"])

    # Time series (bucketed per second)
    buckets: Dict[int, List[Sample]] = defaultdict(list)
    for s in samples:
        buckets[epoch_ms_to_sec_bucket(s.timestamp_ms)].append(s)

    timeseries = []
    for sec_bucket, bucket_samples in sorted(buckets.items()):
        rps = len(bucket_samples)
        avg_rt = statistics.mean([b.elapsed_ms for b in bucket_samples])
        active_threads = bucket_samples[0].all_threads if bucket_samples[0].all_threads is not None else None
        errors_sec = sum(1 for b in bucket_samples if not b.success)  # errors per second

        timeseries.append({
            "sec": sec_bucket,
            "rps": rps,
            "avg_ms": avg_rt,
            "active_threads": active_threads,
            "errors": errors_sec,
        })

    peak_rps = max(ts["rps"] for ts in timeseries) if timeseries else 0

    # Error summary (per response code)
    error_counter = Counter()
    # NEW: detailed error counter per (label, response_code, response_message)
    error_details_counter = Counter()
    for s in samples:
        if not s.success:
            code = s.response_code or "UNKNOWN"
            error_counter[code] += 1
            key = (s.label, code, s.response_message or "")
            error_details_counter[key] += 1

    error_summary = []
    for code, count in error_counter.most_common():
        error_summary.append({
            "response_code": code,
            "count": count,
            "pct": count / errors if errors else 0.0,
        })

    # NEW: error details list
    error_details = []
    for (label, code, msg), count in error_details_counter.most_common():
        error_details.append({
            "label": label,
            "response_code": code,
            "response_message": msg,
            "count": count,
            "pct": count / errors if errors else 0.0,
        })

    metrics = {
        "global": global_stats,
        "per_label": per_label_summaries,
        "timeseries": timeseries,
        "peak_rps": peak_rps,
        "error_summary": error_summary,
        "error_details": error_details,  # NEW
        "start_time": min_ts,
        "end_time": max_ts,
    }
    return metrics


# =========================
# SLA EVALUATION
# =========================

def get_label_sla(label: str, sla_config: Dict[str, Dict[str, float]]) -> Dict[str, float]:
    if label in sla_config:
        base = dict(sla_config.get("*", {}))
        base.update(sla_config[label])
        return base
    return sla_config.get("*", {})


def evaluate_sla(metrics: Dict[str, Any],
                 sla_config: Dict[str, Dict[str, float]]) -> Dict[str, Any]:
    per_label = metrics.get("per_label", [])
    duration_sec = metrics.get("global", {}).get("duration_sec", 0.0)
    rows = []

    for ls in per_label:
        label = ls["label"]
        sla = get_label_sla(label, sla_config)
        if not sla:
            continue

        row = {"label": label, "checks": []}

        if "p95_ms" in sla and ls["p95"] is not None:
            target = sla["p95_ms"]
            actual = ls["p95"]
            status = actual <= target
            row["checks"].append({
                "metric": "P95",
                "target": target,
                "actual": actual,
                "status": status,
            })

        if "avg_ms" in sla and ls["avg"] is not None:
            target = sla["avg_ms"]
            actual = ls["avg"]
            status = actual <= target
            row["checks"].append({
                "metric": "Avg",
                "target": target,
                "actual": actual,
                "status": status,
            })

        if "error_rate" in sla:
            target = sla["error_rate"]
            actual = ls["error_rate"]
            status = actual <= target
            row["checks"].append({
                "metric": "Error Rate",
                "target": target,
                "actual": actual,
                "status": status,
            })

        # Optional TPS SLA
        if "tps_min" in sla and duration_sec > 0:
            target = sla["tps_min"]
            actual = ls["count"] / duration_sec
            status = actual >= target
            row["checks"].append({
                "metric": "TPS",
                "target": target,
                "actual": actual,
                "status": status,
            })

        if row["checks"]:
            rows.append(row)

    total_checks = sum(len(r["checks"]) for r in rows)
    passed_checks = sum(1 for r in rows for c in r["checks"] if c["status"])
    overall_status = (passed_checks == total_checks) if total_checks else True
    overall_pct = (passed_checks / total_checks) if total_checks else 1.0

    return {
        "rows": rows,
        "total_checks": total_checks,
        "passed_checks": passed_checks,
        "overall_status": overall_status,
        "overall_pct": overall_pct,
    }


# =========================
# BASELINE COMPARISON
# =========================

def compare_baseline(current: Dict[str, Any],
                     baseline: Dict[str, Any]) -> Dict[str, Any]:
    cur_map = {x["label"]: x for x in current.get("per_label", [])}
    base_map = {x["label"]: x for x in baseline.get("per_label", [])}
    labels = sorted(set(cur_map.keys()) | set(base_map.keys()))
    rows = []
    for label in labels:
        cur = cur_map.get(label)
        base = base_map.get(label)
        rows.append({
            "label": label,
            "baseline_p95": base["p95"] if base else None,
            "current_p95": cur["p95"] if cur else None,
            "delta_p95": (cur["p95"] - base["p95"])
            if (cur and base and base["p95"] is not None) else None,
            "baseline_error": base["error_rate"] if base else None,
            "current_error": cur["error_rate"] if cur else None,
            "delta_error": (cur["error_rate"] - base["error_rate"])
            if (cur and base) else None,
        })
    return {"rows": rows}


# =========================
# ENHANCED INSIGHTS ENGINE v3.9
# =========================

def calculate_health_score(global_stats: Dict[str, Any],
                           sla_result: Dict[str, Any],
                           metrics: Dict[str, Any]) -> Dict[str, Any]:
    """
    Calculate a composite health score (0-100) with breakdown.
    """
    # SLA Compliance (0-40 points)
    sla_score = 40 * sla_result["overall_pct"] if sla_result["total_checks"] > 0 else 40

    # Error Rate (0-25 points)
    error_rate = global_stats["error_rate"]
    if error_rate <= 0.01:  # 1% or less
        error_score = 25
    elif error_rate <= 0.05:  # 5% or less
        error_score = 20
    elif error_rate <= 0.10:  # 10% or less
        error_score = 10
    else:
        error_score = 0

    # Latency Stability (0-20 points)
    timeseries = metrics.get("timeseries", [])
    if len(timeseries) >= 10:
        early_latency = statistics.mean([ts["avg_ms"] for ts in timeseries[:5]])
        late_latency = statistics.mean([ts["avg_ms"] for ts in timeseries[-5:]])
        latency_ratio = late_latency / early_latency if early_latency > 0 else 1.0

        if latency_ratio <= 1.2:
            latency_score = 20
        elif latency_ratio <= 1.5:
            latency_score = 15
        elif latency_ratio <= 2.0:
            latency_score = 5
        else:
            latency_score = 0
    else:
        latency_score = 10  # Default if insufficient data

    # Throughput Consistency (0-15 points)
    if len(timeseries) >= 10:
        throughput_values = [ts["rps"] for ts in timeseries]
        throughput_cv = statistics.stdev(throughput_values) / statistics.mean(throughput_values) if statistics.mean(
            throughput_values) > 0 else 1.0

        if throughput_cv <= 0.2:
            throughput_score = 15
        elif throughput_cv <= 0.4:
            throughput_score = 10
        elif throughput_cv <= 0.6:
            throughput_score = 5
        else:
            throughput_score = 0
    else:
        throughput_score = 7  # Default

    total_score = sla_score + error_score + latency_score + throughput_score

    # Determine letter grade
    if total_score >= 90:
        grade = "A"
        grade_color = "#16a34a"
    elif total_score >= 80:
        grade = "B"
        grade_color = "#059669"
    elif total_score >= 70:
        grade = "C"
        grade_color = "#ca8a04"
    elif total_score >= 60:
        grade = "D"
        grade_color = "#ea580c"
    else:
        grade = "F"
        grade_color = "#dc2626"

    return {
        "overall": total_score,
        "components": {
            "sla": sla_score,
            "errors": error_score,
            "latency_stability": latency_score,
            "throughput_consistency": throughput_score
        },
        "grade": grade,
        "grade_color": grade_color,
        "interpretation": "Excellent" if total_score >= 90 else
        "Good" if total_score >= 80 else
        "Fair" if total_score >= 70 else
        "Poor" if total_score >= 60 else
        "Critical"
    }


def analyze_root_causes(metrics: Dict[str, Any],
                        sla_result: Dict[str, Any],
                        infra_metrics: Optional[Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    """
    Analyze metrics to suggest probable root causes.
    Returns list of causes with confidence levels.
    """
    causes = []
    global_stats = metrics["global"]
    per_label = metrics["per_label"]
    timeseries = metrics.get("timeseries", [])

    # Check for database bottlenecks
    high_p95_labels = [ls for ls in per_label if ls.get("p95", 0) > 3000]  # >3s P95
    if high_p95_labels and global_stats.get("avg_ms", 0) < 1000:  # High P95 but low average
        causes.append({
            "title": "Database/External Dependency Bottleneck",
            "description": "High P95 latency with relatively low average suggests slow database queries or external API calls affecting some requests.",
            "confidence": "high",
            "confidence_color": "#dc2626",
            "affected_transactions": [ls["label"] for ls in high_p95_labels[:3]],
            "remediation": "Review database query plans, add indexes, optimize joins, or implement caching for slow queries.",
            "icon": "🗄️"
        })

    # Check for thread contention
    if len(timeseries) > 10:
        throughput_trend = calculate_trend([ts["rps"] for ts in timeseries])
        latency_trend = calculate_trend([ts["avg_ms"] for ts in timeseries])

        if latency_trend > 1.5 and throughput_trend < 0.8:  # Latency up, throughput down
            causes.append({
                "title": "Thread Contention or Resource Saturation",
                "description": "Increasing response times with decreasing throughput suggests thread pool exhaustion or resource contention.",
                "confidence": "medium",
                "confidence_color": "#ea580c",
                "remediation": "Increase thread pool size, optimize connection pooling, or scale horizontally.",
                "icon": "⚙️"
            })

    # Check for memory pressure (if infra metrics available)
    if infra_metrics:
        load_section = infra_metrics.get("Load Test") or infra_metrics.get("Soak Test") or {}
        memory_data = load_section.get("Memory")

        if memory_data and memory_data.get("series"):
            memory_values = []
            for series in memory_data["series"]:
                memory_values.extend([v for v in series.get("values", []) if v is not None])

            if memory_values and max(memory_values) > 85:  # Memory >85%
                causes.append({
                    "title": "Memory Pressure Detected",
                    "description": f"Memory utilization reached {max(memory_values):.1f}%, potentially causing garbage collection overhead.",
                    "confidence": "high",
                    "confidence_color": "#dc2626",
                    "remediation": "Increase JVM heap size, optimize memory usage, or scale memory resources.",
                    "icon": "🧠"
                })

    # Check for error patterns
    error_summary = metrics.get("error_summary", [])
    if error_summary:
        top_error = error_summary[0]
        if top_error["response_code"] in ["500", "502", "503", "504"]:
            causes.append({
                "title": f"Server-side Errors ({top_error['response_code']}) Dominating",
                "description": f"{top_error['count']} occurrences of HTTP {top_error['response_code']} errors suggest backend service instability.",
                "confidence": "high",
                "confidence_color": "#dc2626",
                "remediation": "Check application server logs, review recent deployments, and validate database connections.",
                "icon": "🚨"
            })

    # Check for network issues (high latency with low CPU)
    if infra_metrics:
        load_section = infra_metrics.get("Load Test") or infra_metrics.get("Soak Test") or {}
        cpu_data = load_section.get("CPU")

        if cpu_data and cpu_data.get("series") and global_stats.get("p95_ms", 0) > 2000:
            cpu_values = []
            for series in cpu_data["series"]:
                cpu_values.extend([v for v in series.get("values", []) if v is not None])

            if cpu_values and statistics.mean(cpu_values) < 50:  # Low CPU but high latency
                causes.append({
                    "title": "Possible Network or I/O Bottleneck",
                    "description": "High latency with low CPU utilization suggests network delays, disk I/O issues, or external service dependencies.",
                    "confidence": "medium",
                    "confidence_color": "#ca8a04",
                    "remediation": "Check network latency, review external service SLAs, and monitor disk I/O metrics.",
                    "icon": "🌐"
                })

    # Check for cache effectiveness
    if len(timeseries) > 20:
        early_latency = statistics.mean([ts["avg_ms"] for ts in timeseries[:10]])
        late_latency = statistics.mean([ts["avg_ms"] for ts in timeseries[-10:]])

        if late_latency < early_latency * 0.7:  # 30% improvement
            causes.append({
                "title": "Cache Warming Observed",
                "description": "Response times improved significantly over test duration, indicating cache warming effect.",
                "confidence": "low",
                "confidence_color": "#16a34a",
                "remediation": "Consider implementing pre-warming for production deployments.",
                "icon": "🔥"
            })

    return causes


def estimate_capacity(metrics: Dict[str, Any]) -> Dict[str, Any]:
    """
    Estimate system capacity based on current performance.
    """
    global_stats = metrics["global"]
    peak_rps = metrics["peak_rps"]
    avg_rt = global_stats["avg_ms"]
    p95_rt = global_stats["p95_ms"]
    error_rate = global_stats["error_rate"]

    # Simple capacity estimation (using Little's Law)
    sustainable_rps = peak_rps * 0.8  # Assume 80% of peak is sustainable

    # Estimate concurrent users (using average think time of 5 seconds)
    think_time_ms = 5000
    concurrent_users = sustainable_rps * ((avg_rt + think_time_ms) / 1000)

    # Identify bottlenecks
    bottlenecks = []
    if p95_rt > 3000:
        bottlenecks.append({"issue": "High latency (P95 > 3s)", "severity": "high"})

    if error_rate > 0.05:
        bottlenecks.append({"issue": f"High error rate ({error_rate * 100:.1f}%)", "severity": "high"})

    if avg_rt > 1000:
        bottlenecks.append({"issue": "Slow average response (>1s)", "severity": "medium"})

    # Capacity grade
    if len([b for b in bottlenecks if b["severity"] == "high"]) > 0:
        capacity_grade = "Limited"
        capacity_color = "#dc2626"
    elif len(bottlenecks) > 0:
        capacity_grade = "Moderate"
        capacity_color = "#ca8a04"
    else:
        capacity_grade = "Good"
        capacity_color = "#16a34a"

    return {
        "estimated_sustainable_rps": sustainable_rps,
        "estimated_concurrent_users": int(concurrent_users),
        "peak_handled": peak_rps,
        "bottlenecks": bottlenecks,
        "capacity_grade": capacity_grade,
        "capacity_color": capacity_color,
        "recommended_action": "Scale horizontally and optimize bottlenecks" if len(
            bottlenecks) > 0 else "Current capacity appears adequate"
    }


def generate_priority_recommendations(global_stats: Dict[str, Any],
                                      sla_result: Dict[str, Any],
                                      metrics: Dict[str, Any],
                                      baseline_comparison: Optional[Dict[str, Any]] = None,
                                      infra_metrics: Optional[Dict[str, Any]] = None) -> Dict[
    str, List[Dict[str, Any]]]:
    """
    Generate prioritized recommendations with action items.
    """
    recommendations = {
        "critical": [],
        "high": [],
        "medium": [],
        "low": []
    }

    # CRITICAL: SLA failures
    if not sla_result["overall_status"]:
        failing_transactions = []
        for row in sla_result.get("rows", []):
            for check in row.get("checks", []):
                if not check["status"]:
                    failing_transactions.append(f"{row['label']} ({check['metric']})")

        recommendations["critical"].append({
            "title": "SLA Compliance Failures",
            "description": f"{sla_result['passed_checks']}/{sla_result['total_checks']} checks passed.",
            "action_items": [
                f"Investigate failing transactions: {', '.join(failing_transactions[:3])}",
                "Review and adjust SLA thresholds if appropriate",
                "Prioritize fixes for transactions with highest breach severity"
            ],
            "metrics": {
                "passed_checks": sla_result["passed_checks"],
                "total_checks": sla_result["total_checks"],
                "pass_rate": f"{sla_result['overall_pct'] * 100:.1f}%"
            },
            "icon": "🚨"
        })

    # CRITICAL: High error rate
    error_rate = global_stats["error_rate"]
    if error_rate > 0.10:  # >10% errors
        recommendations["critical"].append({
            "title": "Excessive Error Rate",
            "description": f"Error rate of {error_rate * 100:.1f}% exceeds acceptable thresholds.",
            "action_items": [
                "Review error logs and patterns immediately",
                "Check application and database health",
                "Consider rolling back recent changes if errors spiked"
            ],
            "icon": "💥"
        })

    # HIGH: Performance degradation vs baseline
    if baseline_comparison and baseline_comparison.get("rows"):
        regressions = []
        severe_regressions = []
        for row in baseline_comparison["rows"]:
            if row.get("delta_p95", 0) > 500:  # >500ms regression
                regressions.append(f"{row['label']} (+{row['delta_p95']:.0f}ms)")
                if row.get("delta_p95", 0) > 1000:  # >1s regression
                    severe_regressions.append(row['label'])

        if regressions:
            severity = "high" if severe_regressions else "medium"
            recommendations[severity].append({
                "title": "Performance Regressions Detected",
                "description": f"{len(regressions)} transactions show P95 latency increase >500ms.",
                "action_items": [
                    f"Investigate regressed transactions: {', '.join(regressions[:3])}",
                    "Compare configurations between current and baseline runs",
                    "Review code changes in affected areas"
                ],
                "icon": "📉"
            })

    # HIGH: Resource saturation
    if infra_metrics:
        load_section = infra_metrics.get("Load Test") or infra_metrics.get("Soak Test") or {}
        cpu_data = load_section.get("CPU")

        if cpu_data and cpu_data.get("series"):
            cpu_values = []
            for series in cpu_data["series"]:
                cpu_values.extend([v for v in series.get("values", []) if v is not None])

            if cpu_values and max(cpu_values) > 85:
                recommendations["high"].append({
                    "title": "CPU Saturation Detected",
                    "description": f"CPU utilization reached {max(cpu_values):.1f}% during test.",
                    "action_items": [
                        "Consider horizontal scaling (add more instances)",
                        "Optimize CPU-intensive operations",
                        "Review application profiling data"
                    ],
                    "icon": "🔥"
                })

    # MEDIUM: Latency stability issues
    timeseries = metrics.get("timeseries", [])
    if len(timeseries) >= 20:
        segment_size = len(timeseries) // 4
        segments = [
            timeseries[:segment_size],
            timeseries[segment_size:2 * segment_size],
            timeseries[2 * segment_size:3 * segment_size],
            timeseries[3 * segment_size:]
        ]

        segment_avgs = [statistics.mean([s["avg_ms"] for s in seg]) for seg in segments]
        if segment_avgs[-1] > segment_avgs[0] * 1.3:  # 30% increase
            recommendations["medium"].append({
                "title": "Gradual Latency Increase",
                "description": "Response time increased by >30% from start to end of test.",
                "action_items": [
                    "Check for memory leaks or resource exhaustion",
                    "Monitor database connection pool usage",
                    "Consider implementing connection recycling"
                ],
                "icon": "📈"
            })

    # MEDIUM: High variance in response times
    per_label = metrics.get("per_label", [])
    high_variance_labels = []
    for ls in per_label:
        if ls.get("p95", 0) > 0 and ls.get("p50", 0) > 0:
            variance_ratio = ls["p95"] / ls["p50"] if ls["p50"] > 0 else 1.0
            if variance_ratio > 3.0:  # P95 is 3x P50
                high_variance_labels.append(ls["label"])

    if high_variance_labels:
        recommendations["medium"].append({
            "title": "High Response Time Variance",
            "description": f"{len(high_variance_labels)} transactions show high variance between P50 and P95.",
            "action_items": [
                f"Investigate inconsistent performance: {', '.join(high_variance_labels[:3])}",
                "Check for uneven load distribution",
                "Review caching effectiveness"
            ],
            "icon": "🎢"
        })

    # LOW: Optimization opportunities
    low_tps_transactions = []
    for ls in metrics.get("per_label", []):
        if ls.get("count", 0) > 10:  # Minimum samples
            tps = ls["count"] / global_stats["duration_sec"] if global_stats["duration_sec"] > 0 else 0
            if tps < 1.0:  # Very low throughput
                low_tps_transactions.append(ls["label"])

    if low_tps_transactions:
        recommendations["low"].append({
            "title": "Low Throughput Transactions",
            "description": f"{len(low_tps_transactions)} transactions have very low throughput (<1 TPS).",
            "action_items": [
                f"Review transaction design: {', '.join(low_tps_transactions[:3])}",
                "Consider batching or async processing",
                "Evaluate if these transactions are necessary for load test"
            ],
            "icon": "🐢"
        })

    # LOW: Good performance recognition
    if sla_result["overall_status"] and error_rate < 0.01 and len(recommendations["critical"]) == 0:
        recommendations["low"].append({
            "title": "Excellent Performance",
            "description": "All SLAs met with minimal errors. System is performing well under load.",
            "action_items": [
                "Consider increasing load to find next breaking point",
                "Document configuration as baseline for future tests",
                "Proceed with confidence to next environment"
            ],
            "icon": "✅"
        })

    return recommendations


# =========================
# CONTAINER COMPARISON ANALYSIS (NEW in v3.9)
# =========================

def analyze_container_comparison(infra_data: Dict[str, Any]) -> Dict[str, Any]:
    """
    Analyze container resource trends based on infrastructure data timeline.
    Uses first 2 minutes vs last 2 minutes of available infrastructure data.
    """
    if not infra_data:
        return {}

    results = {
        "containers": {},
        "summary": {
            "cpu_increase_count": 0,
            "memory_increase_count": 0,
            "high_increase_containers": [],
            "analysis_method": "infra_timeline",
            "data_timeline": {
                "total_minutes": 0,
                "before_minutes": 0,
                "after_minutes": 0
            }
        }
    }

    # Find the active section (Load Test or Soak Test)
    load_section = infra_data.get("Load Test") or {}
    soak_section = infra_data.get("Soak Test") or {}
    active_section = load_section if load_section else soak_section

    if not active_section:
        return results

    # Analyze CPU data
    cpu_data = active_section.get("CPU")
    mem_data = active_section.get("Memory")

    if cpu_data and cpu_data.get("labels") and cpu_data.get("series"):
        labels = cpu_data["labels"]

        # Parse timestamps and create relative timeline
        label_times = []
        for label in labels:
            # Try multiple date formats
            for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%H:%M:%S", "%Y/%m/%d %H:%M:%S"]:
                try:
                    dt = datetime.strptime(label, fmt)
                    label_times.append(int(dt.timestamp()))
                    break
                except ValueError:
                    continue
            else:
                # If no format worked, create relative timestamps starting from 0
                label_times.append(len(label_times) * 60)  # 1 minute intervals

        if not label_times:
            return results

        # Determine data duration
        min_time = min(label_times)
        max_time = max(label_times)
        data_duration_sec = max_time - min_time

        # Define analysis periods (first 2 min vs last 2 min of infrastructure data)
        if data_duration_sec > 240:  # More than 4 minutes of data
            before_end_time = min_time + 120  # First 2 minutes
            after_start_time = max_time - 120  # Last 2 minutes
            before_minutes = 2
            after_minutes = 2
        else:
            # For shorter data, use first 1/3 vs last 1/3
            segment = data_duration_sec // 3
            before_end_time = min_time + segment
            after_start_time = max_time - segment
            before_minutes = segment / 60
            after_minutes = segment / 60

        results["summary"]["data_timeline"]["total_minutes"] = data_duration_sec / 60
        results["summary"]["data_timeline"]["before_minutes"] = before_minutes
        results["summary"]["data_timeline"]["after_minutes"] = after_minutes

        # Process each container series
        for series in cpu_data["series"]:
            container = series.get("name", "Unknown")
            values = series.get("values", [])

            if not values or len(values) < 2:
                continue

            # Find indices for before (first segment) and after (last segment) periods
            before_indices = []
            after_indices = []

            for i, (label_time, value) in enumerate(zip(label_times, values)):
                if value is None:
                    continue

                # First segment of infrastructure data
                if min_time <= label_time <= before_end_time:
                    before_indices.append((i, value))
                # Last segment of infrastructure data
                elif after_start_time <= label_time <= max_time:
                    after_indices.append((i, value))

            # Need at least 2 data points in each period for meaningful comparison
            if len(before_indices) < 2 or len(after_indices) < 2:
                continue

            # Calculate averages
            before_avg = statistics.mean([v for _, v in before_indices])
            after_avg = statistics.mean([v for _, v in after_indices])

            # Store results
            if container not in results["containers"]:
                results["containers"][container] = {}

            increase_pct = None
            if before_avg > 0:
                increase_pct = ((after_avg - before_avg) / before_avg * 100)

            results["containers"][container]["cpu"] = {
                "before": before_avg,
                "after": after_avg,
                "increase_pct": increase_pct,
                "has_before": True,
                "has_after": True,
                "data_points_before": len(before_indices),
                "data_points_after": len(after_indices),
                "min_time": min_time,
                "max_time": max_time
            }

            # Track significant increases
            if increase_pct and increase_pct > 50:
                results["summary"]["cpu_increase_count"] += 1
                if container not in results["summary"]["high_increase_containers"]:
                    results["summary"]["high_increase_containers"].append(container)

    # Analyze Memory data
    if mem_data and mem_data.get("labels") and mem_data.get("series"):
        labels = mem_data["labels"]
        label_times = []
        for label in labels:
            for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%H:%M:%S", "%Y/%m/%d %H:%M:%S"]:
                try:
                    dt = datetime.strptime(label, fmt)
                    label_times.append(int(dt.timestamp()))
                    break
                except ValueError:
                    continue
            else:
                label_times.append(len(label_times) * 60)

        if label_times:
            min_time = min(label_times)
            max_time = max(label_times)
            data_duration_sec = max_time - min_time

            if data_duration_sec > 240:
                before_end_time = min_time + 120
                after_start_time = max_time - 120
            else:
                segment = data_duration_sec // 3
                before_end_time = min_time + segment
                after_start_time = max_time - segment

        for series in mem_data["series"]:
            container = series.get("name", "Unknown")
            values = series.get("values", [])

            if not values or len(values) < 2:
                continue

            before_indices = []
            after_indices = []

            for i, (label_time, value) in enumerate(zip(label_times, values)):
                if value is None:
                    continue

                if min_time <= label_time <= before_end_time:
                    before_indices.append((i, value))
                elif after_start_time <= label_time <= max_time:
                    after_indices.append((i, value))

            if len(before_indices) < 2 or len(after_indices) < 2:
                continue

            before_avg = statistics.mean([v for _, v in before_indices])
            after_avg = statistics.mean([v for _, v in after_indices])

            if container not in results["containers"]:
                results["containers"][container] = {}

            increase_pct = None
            if before_avg > 0:
                increase_pct = ((after_avg - before_avg) / before_avg * 100)

            results["containers"][container]["memory"] = {
                "before": before_avg,
                "after": after_avg,
                "increase_pct": increase_pct,
                "has_before": True,
                "has_after": True,
                "data_points_before": len(before_indices),
                "data_points_after": len(after_indices)
            }

            if increase_pct and increase_pct > 50:
                results["summary"]["memory_increase_count"] += 1
                if container not in results["summary"]["high_increase_containers"]:
                    results["summary"]["high_increase_containers"].append(container)

    return results

#==============================
#build_container_comparison_html
#===============================

def build_container_comparison_html(container_comparison: Dict[str, Any]) -> str:
    """
    Build HTML for container trend analysis based on infrastructure data timeline.
    """
    if not container_comparison or not container_comparison.get("containers"):
        return """
        <div style='padding: 12px; background-color: #f3f4f6; border-radius: 8px; margin-top: 12px;'>
            <div style='font-weight: 600; color: #4b5563;'>No Container Trend Analysis Data</div>
            <div style='font-size: 12px; color: #6b7280; margin-top: 4px;'>
                Trend analysis requires infrastructure metrics with sufficient time-series data.
            </div>
        </div>
        """

    containers = container_comparison["containers"]
    summary = container_comparison.get("summary", {})
    timeline = summary.get("data_timeline", {})

    total_minutes = timeline.get("total_minutes", 0)
    before_minutes = timeline.get("before_minutes", 0)
    after_minutes = timeline.get("after_minutes", 0)

    before_period_str = f"First {before_minutes:.1f} min"
    after_period_str = f"Last {after_minutes:.1f} min"

    items = []
    items.append(f"""
    <div style='font-weight: 600; color: #111827; margin-bottom: 12px;'>
        📈 Container Resource Trend Analysis ({before_period_str} vs {after_period_str})
    </div>
    """)

    # Explanation of the analysis
    items.append(f"""
    <div style="padding: 10px; background-color: #dbeafe; border-radius: 6px; margin-bottom: 12px;">
        <div style="display: flex; align-items: center; gap: 8px; color: #1e40af;">
            <span>📊</span>
            <div style="font-weight: 600;">Analysis Based on Infrastructure Timeline</div>
        </div>
        <div style="font-size: 12px; color: #374151; margin-top: 4px;">
            Comparing resource usage between <strong>{before_period_str}</strong> and 
            <strong>{after_period_str}</strong> of {total_minutes:.1f} minutes infrastructure data.
            Shows resource trends during the monitoring period.
        </div>
    </div>
    """)

    # Summary card
    cpu_increase = summary.get('cpu_increase_count', 0) or 0
    mem_increase = summary.get('memory_increase_count', 0) or 0

    items.append(f"""
    <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 12px; margin-bottom: 16px;">
        <div style="padding: 12px; background: #dbeafe; border-radius: 8px;">
            <div style="font-size: 12px; color: #1e40af; margin-bottom: 4px;">Containers Analyzed</div>
            <div style="font-size: 20px; font-weight: 700; color: #1e40af;">{len(containers)}</div>
        </div>
        <div style="padding: 12px; background: #fef3c7; border-radius: 8px;">
            <div style="font-size: 12px; color: #92400e; margin-bottom: 4px;">CPU Trend Up (>50%)</div>
            <div style="font-size: 20px; font-weight: 700; color: #92400e;">{cpu_increase}</div>
        </div>
        <div style="padding: 12px; background: #fce7f3; border-radius: 8px;">
            <div style="font-size: 12px; color: #9d174d; margin-bottom: 4px;">Memory Trend Up (>50%)</div>
            <div style="font-size: 20px; font-weight: 700; color: #9d174d;">{mem_increase}</div>
        </div>
    </div>
    """)

    # Data quality indicator
    containers_with_sufficient_data = sum(1 for c in containers.values()
                                          if c.get("cpu", {}).get("data_points_before", 0) >= 2 and
                                          c.get("cpu", {}).get("data_points_after", 0) >= 2)

    items.append(f"""
    <div style="font-size: 11px; color: #6b7280; padding: 8px; background: #f9fafb; border-radius: 4px; margin-bottom: 12px;">
        <div style="display: flex; gap: 16px; flex-wrap: wrap;">
            <div>Infrastructure Data: {total_minutes:.1f} minutes total</div>
            <div>Analysis Periods: {before_period_str} • {after_period_str}</div>
            <div>Quality: {containers_with_sufficient_data}/{len(containers)} containers with sufficient data</div>
        </div>
    </div>
    """)

    # Container table with more precision
    items.append("""
    <div style="max-height: 400px; overflow-y: auto; border: 1px solid #e5e7eb; border-radius: 6px;">
        <table style="width: 100%; border-collapse: collapse; font-size: 12px;">
            <thead style="background: #f8fafc; position: sticky; top: 0;">
                <tr>
                    <th style="padding: 10px; text-align: left; border-bottom: 1px solid #e5e7eb;">Container</th>
                    <th style="padding: 10px; text-align: left; border-bottom: 1px solid #e5e7eb;">CPU Early</th>
                    <th style="padding: 10px; text-align: left; border-bottom: 1px solid #e5e7eb;">CPU Late</th>
                    <th style="padding: 10px; text-align: left; border-bottom: 1px solid #e5e7eb;">Trend</th>
                    <th style="padding: 10px; text-align: left; border-bottom: 1px solid #e5e7eb;">Memory Early</th>
                    <th style="padding: 10px; text-align: left; border-bottom: 1px solid #e5e7eb;">Memory Late</th>
                    <th style="padding: 10px; text-align: left; border-bottom: 1px solid #e5e7eb;">Trend</th>
                </tr>
            </thead>
            <tbody>
    """)

    for container_name, data in sorted(containers.items()):
        cpu = data.get("cpu", {})
        mem = data.get("memory", {})

        # Check if we have sufficient data
        cpu_has_data = cpu.get("has_before", False) and cpu.get("has_after", False)
        mem_has_data = mem.get("has_before", False) and mem.get("has_after", False)

        # Get raw values for debugging
        cpu_early_raw = cpu.get('before')
        cpu_late_raw = cpu.get('after')
        cpu_trend_raw = cpu.get('increase_pct')

        mem_early_raw = mem.get('before')
        mem_late_raw = mem.get('after')
        mem_trend_raw = mem.get('increase_pct')

        # Format values with more precision for low values
        def format_value(value, default="-"):
            if value is None:
                return f'<span style="color: #9ca3af;">{default}</span>'
            try:
                val = float(value)
                # Use more precision for values < 1%
                if val < 1.0:
                    return f"{val:.2f}%"
                else:
                    return f"{val:.1f}%"
            except (TypeError, ValueError):
                return f'<span style="color: #9ca3af;">{default}</span>'

        # Format values with actual precision used in calculation
        cpu_early = format_value(cpu_early_raw)
        cpu_late = format_value(cpu_late_raw)

        mem_early = format_value(mem_early_raw)
        mem_late = format_value(mem_late_raw)

        # Determine trend display - FIXED to use actual calculation
        def get_trend_display(pct, early_val, late_val, has_data=True):
            if not has_data or pct is None or early_val is None or late_val is None:
                return f'<span style="color: #9ca3af;">-</span>'

            try:
                pct_float = float(pct)
                early_float = float(early_val)
                late_float = float(late_val)

                # For debugging - show actual values
                debug_info = ""
                # Uncomment for debugging: debug_info = f' ({early_float:.3f}→{late_float:.3f})'

                # Calculate actual percentage change to verify
                actual_pct = ((late_float - early_float) / early_float * 100) if early_float > 0 else 0

                # Use consistent threshold logic
                if pct_float > 100:
                    return f'<span style="color: #dc2626; font-weight: bold;" title="Early: {early_float:.3f}%, Late: {late_float:.3f}%">↗ {pct_float:.0f}%{debug_info}</span>'
                elif pct_float > 50:
                    return f'<span style="color: #dc2626;" title="Early: {early_float:.3f}%, Late: {late_float:.3f}%">↗ {pct_float:.0f}%{debug_info}</span>'
                elif pct_float > 10:
                    return f'<span style="color: #ea580c;" title="Early: {early_float:.3f}%, Late: {late_float:.3f}%">↗ {pct_float:.0f}%{debug_info}</span>'
                elif pct_float < -10:
                    return f'<span style="color: #16a34a;" title="Early: {early_float:.3f}%, Late: {late_float:.3f}%">↘ {abs(pct_float):.0f}%{debug_info}</span>'
                else:
                    return f'<span style="color: #6b7280;" title="Early: {early_float:.3f}%, Late: {late_float:.3f}%">→ {pct_float:.0f}%{debug_info}</span>'
            except:
                return f'<span style="color: #9ca3af;">-</span>'

        items.append(f"""
        <tr style="border-bottom: 1px solid #f3f4f6;">
            <td style="padding: 10px; font-family: ui-monospace; font-size: 11px;">{container_name}</td>
            <td style="padding: 10px;">{cpu_early}</td>
            <td style="padding: 10px; font-weight: 500;">{cpu_late}</td>
            <td style="padding: 10px;">{get_trend_display(cpu_trend_raw, cpu_early_raw, cpu_late_raw, cpu_has_data)}</td>
            <td style="padding: 10px;">{mem_early}</td>
            <td style="padding: 10px; font-weight: 500;">{mem_late}</td>
            <td style="padding: 10px;">{get_trend_display(mem_trend_raw, mem_early_raw, mem_late_raw, mem_has_data)}</td>
        </tr>
        """)

    items.append("""
            </tbody>
        </table>
    </div>

    <div style="margin-top: 12px; font-size: 11px; color: #6b7280; padding: 8px; background: #f9fafb; border-radius: 4px;">
        <div style="display: flex; gap: 12px; flex-wrap: wrap;">
            <div style="display: flex; align-items: center; gap: 4px;">
                <span style="color: #16a34a; font-weight: bold;">↘</span>
                <span>Decreased (>10%)</span>
            </div>
            <div style="display: flex; align-items: center; gap: 4px;">
                <span style="color: #6b7280;">→</span>
                <span>Stable (±10%)</span>
            </div>
            <div style="display: flex; align-items: center; gap: 4px;">
                <span style="color: #ea580c;">↗</span>
                <span>Moderate increase (10-50%)</span>
            </div>
            <div style="display: flex; align-items: center; gap: 4px;">
                <span style="color: #dc2626; font-weight: bold;">↗</span>
                <span>High increase (>50%)</span>
            </div>
            <div style="display: flex; align-items: center; gap: 4px; margin-left: 8px;">
                <small>Hover over trend for exact values</small>
            </div>
        </div>
    </div>
    """)

    # Add explanation for small percentage changes
    items.append("""
    <div style="margin-top: 16px; padding: 12px; background-color: #f3f4f6; border-radius: 6px;">
        <div style="font-weight: 600; color: #4b5563; margin-bottom: 4px;">📝 Understanding Small Percentage Changes</div>
        <div style="font-size: 12px; color: #6b7280;">
            <strong>Why small values show large percentage changes:</strong>
            <div style="padding-left: 8px; margin-top: 4px;">
                • Example: CPU changes from 0.21% to 0.17% = -19% change<br>
                • Both values round to 0.2% display, but actual change is significant<br>
                • For low baseline values, small absolute changes create large percentage changes<br>
                • This highlights sensitivity to small variations in low-utilization containers
            </div>
        </div>
    </div>
    """)

    return "<div style='padding: 8px;'>" + "\n".join(items) + "</div>"

# =========================
# EXCEL EXPORT FUNCTIONALITY (NEW in v3.7)
# =========================

def export_to_excel(metrics: Dict[str, Any],
                    sla_result: Dict[str, Any],
                    container_comparison: Dict[str, Any],
                    test_name: str,
                    environment: str,
                    output_path: str) -> str:
    """
    Export complete report to Excel with rich formatting, charts, and multiple tabs.
    """
    if not EXCEL_SUPPORT:
        raise ImportError("openpyxl is not installed. Install with: pip install openpyxl")

    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        default_sheet = wb['Sheet']
        wb.remove(default_sheet)

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    # 1. EXECUTIVE SUMMARY TAB
    ws_summary = wb.create_sheet(title="Executive Summary")
    ws_summary.sheet_view.showGridLines = False

    # Add logo/header
    ws_summary.merge_cells('A1:H1')
    ws_summary['A1'] = f"JMETER PERFORMANCE TEST REPORT - {test_name}"
    ws_summary['A1'].font = Font(size=16, bold=True, color="366092")
    ws_summary['A1'].alignment = Alignment(horizontal='center')

    ws_summary['A3'] = "Test Information"
    ws_summary['A3'].font = Font(bold=True, size=12)

    test_info = [
        ["Test Name:", test_name],
        ["Environment:", environment],
        ["Start Time:", datetime.fromtimestamp(metrics["start_time"] / 1000).strftime("%Y-%m-%d %H:%M:%S")],
        ["End Time:", datetime.fromtimestamp(metrics["end_time"] / 1000).strftime("%Y-%m-%d %H:%M:%S")],
        ["Duration:", f"{metrics['global']['duration_sec']:.0f} seconds"],
        ["Total Requests:", f"{metrics['global']['total_requests']:,}"],
        ["Error Rate:", f"{metrics['global']['error_rate'] * 100:.2f}%"],
        ["SLA Status:", "PASS" if sla_result["overall_status"] else "FAIL"]
    ]

    for i, (label, value) in enumerate(test_info, start=4):
        ws_summary[f'A{i}'] = label
        ws_summary[f'B{i}'] = value
        ws_summary[f'A{i}'].font = Font(bold=True)

    # Health Score Box
    health_score = calculate_health_score(metrics['global'], sla_result, metrics)
    ws_summary['D4'] = "Health Score"
    ws_summary['D4'].font = Font(bold=True, size=12)
    ws_summary.merge_cells('D5:F7')
    ws_summary['D5'] = f"{health_score['overall']}/100\n{health_score['grade']}"
    ws_summary['D5'].font = Font(size=24, bold=True)
    ws_summary['D5'].alignment = Alignment(horizontal='center', vertical='center')

    # Color fill based on score
    if health_score['overall'] >= 80:
        fill_color = "C6EFCE"  # Green
        text_color = "006100"
    elif health_score['overall'] >= 60:
        fill_color = "FFEB9C"  # Yellow
        text_color = "9C6500"
    else:
        fill_color = "FFC7CE"  # Red
        text_color = "9C0006"

    ws_summary['D5'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    ws_summary['D5'].font = Font(size=24, bold=True, color=text_color)

    # KPI Dashboard
    ws_summary['A10'] = "Key Performance Indicators"
    ws_summary['A10'].font = Font(bold=True, size=12)

    kpis = [
        ["Metric", "Value", "Target", "Status"],
        ["P95 Latency (ms)", f"{metrics['global']['p95_ms']:.1f}", "< 3000",
         "✅" if metrics['global']['p95_ms'] < 3000 else "❌"],
        ["Error Rate", f"{metrics['global']['error_rate'] * 100:.2f}%", "< 2%",
         "✅" if metrics['global']['error_rate'] < 0.02 else "❌"],
        ["Peak Throughput (RPS)", f"{metrics['peak_rps']:.1f}", "> 50",
         "✅" if metrics['peak_rps'] > 50 else "⚠️"],
        ["APDEX Score", f"{metrics['global']['apdex']:.3f}", "> 0.9",
         "✅" if metrics['global']['apdex'] > 0.9 else "❌"],
        ["SLA Compliance", f"{sla_result['passed_checks']}/{sla_result['total_checks']}", "100%",
         "✅" if sla_result['overall_status'] else "❌"]
    ]

    for i, row in enumerate(kpis, start=11):
        for j, value in enumerate(row, start=1):
            cell = ws_summary.cell(row=i, column=j, value=value)
            cell.border = border
            if i == 11:  # Header row
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

    # Set column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws_summary.column_dimensions[col].width = 20

    # 2. TRANSACTION DETAILS TAB
    ws_transactions = wb.create_sheet(title="Transaction Details")

    # Add headers
    headers = ["Transaction", "Count", "Min (ms)", "Avg (ms)", "P50 (ms)",
               "P90 (ms)", "P95 (ms)", "P99 (ms)", "Max (ms)",
               "Throughput (TPS)", "Error Rate", "SLA Status"]

    for col, header in enumerate(headers, start=1):
        cell = ws_transactions.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_alignment

    # Add data rows with conditional formatting
    for i, trans in enumerate(metrics['per_label'], start=2):
        error_rate = trans.get('error_rate', 0)
        p95 = trans.get('p95', 0)
        tps = trans['count'] / metrics['global']['duration_sec'] if metrics['global']['duration_sec'] > 0 else 0

        # Determine SLA status
        sla_status = "PASS"
        status_color = "C6EFCE"  # Green

        if error_rate > 0.02:
            sla_status = "FAIL"
            status_color = "FFC7CE"  # Red
        elif p95 > 3000:
            sla_status = "WARNING"
            status_color = "FFEB9C"  # Yellow

        row_data = [
            trans['label'],
            trans['count'],
            trans.get('min', 0),
            trans.get('avg', 0),
            trans.get('p50', 0),
            trans.get('p90', 0),
            trans.get('p95', 0),
            trans.get('p99', 0),
            trans.get('max', 0),
            tps,
            f"{error_rate * 100:.2f}%",
            sla_status
        ]

        for col, value in enumerate(row_data, start=1):
            cell = ws_transactions.cell(row=i, column=col, value=value)
            cell.border = border

            # Color code based on values
            if col == 11:  # Error Rate column
                if error_rate > 0.05:
                    cell.font = Font(color="FF0000", bold=True)
                elif error_rate > 0.02:
                    cell.font = Font(color="FF9900", bold=True)

            if col == 12:  # SLA Status column
                cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
                cell.font = Font(bold=True)

    # Add auto-filter
    ws_transactions.auto_filter.ref = ws_transactions.dimensions

    # Set column widths
    for col in range(1, len(headers) + 1):
        ws_transactions.column_dimensions[get_column_letter(col)].width = 15

    # 3. TIME SERIES CHARTS TAB
    ws_timeseries = wb.create_sheet(title="Time Series")

    # Add time series data
    timeseries = metrics['timeseries']
    time_headers = ["Timestamp", "RPS", "Avg Response (ms)", "Errors", "Active Threads"]

    for col, header in enumerate(time_headers, start=1):
        cell = ws_timeseries.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for i, ts in enumerate(timeseries, start=2):
        ws_timeseries.cell(row=i, column=1,
                           value=datetime.fromtimestamp(ts['sec']).strftime("%H:%M:%S"))
        ws_timeseries.cell(row=i, column=2, value=ts['rps'])
        ws_timeseries.cell(row=i, column=3, value=ts['avg_ms'])
        ws_timeseries.cell(row=i, column=4, value=ts.get('errors', 0))
        ws_timeseries.cell(row=i, column=5, value=ts.get('active_threads', 0))

    # Create charts
    chart1 = LineChart()
    chart1.title = "Throughput (RPS) Over Time"
    chart1.style = 13
    data = Reference(ws_timeseries, min_col=2, min_row=1, max_row=len(timeseries) + 1)
    cats = Reference(ws_timeseries, min_col=1, min_row=2, max_row=len(timeseries) + 1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.y_axis.title = "Requests Per Second"
    chart1.x_axis.title = "Time"
    ws_timeseries.add_chart(chart1, "G2")

    chart2 = LineChart()
    chart2.title = "Response Time Over Time"
    chart2.style = 13
    data = Reference(ws_timeseries, min_col=3, min_row=1, max_row=len(timeseries) + 1)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.y_axis.title = "Response Time (ms)"
    chart2.x_axis.title = "Time"
    ws_timeseries.add_chart(chart2, "G20")

    # 4. CONTAINER RESOURCES TAB
    if container_comparison and container_comparison.get("containers"):
        ws_containers = wb.create_sheet(title="Container Resources")

        headers = ["Container Name", "CPU Before (%)", "CPU During (%)", "CPU After (%)",
                   "CPU Δ%", "Memory Before (%)", "Memory During (%)", "Memory After (%)",
                   "Memory Δ%", "Status"]

        for col, header in enumerate(headers, start=1):
            cell = ws_containers.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment

        row_num = 2
        for container, data in container_comparison.get("containers", {}).items():
            cpu = data.get("cpu", {})
            mem = data.get("memory", {})

            # Determine overall status
            status = "OK"
            if (cpu.get('increase_pct', 0) > 50 or mem.get('increase_pct', 0) > 50):
                status = "WARNING"
            if (cpu.get('increase_pct', 0) > 100 or mem.get('increase_pct', 0) > 100):
                status = "CRITICAL"

            row_data = [
                container,
                cpu.get('before', 0) if cpu.get('before') is not None else 0,
                cpu.get('during', 0) if cpu.get('during') is not None else 0,
                cpu.get('after', 0) if cpu.get('after') is not None else 0,
                cpu.get('increase_pct', 0) if cpu.get('increase_pct') is not None else 0,
                mem.get('before', 0) if mem.get('before') is not None else 0,
                mem.get('during', 0) if mem.get('during') is not None else 0,
                mem.get('after', 0) if mem.get('after') is not None else 0,
                mem.get('increase_pct', 0) if mem.get('increase_pct') is not None else 0,
                status
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws_containers.cell(row=row_num, column=col, value=value)
                cell.border = border

                # Color code status
                if col == 10:  # Status column
                    if status == "CRITICAL":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006", bold=True)
                    elif status == "WARNING":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(color="9C6500", bold=True)
                    else:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100", bold=True)

            row_num += 1

        # Set column widths
        for col in range(1, len(headers) + 1):
            ws_containers.column_dimensions[get_column_letter(col)].width = 15

    # 5. ERROR ANALYSIS TAB
    ws_errors = wb.create_sheet(title="Error Analysis")

    error_headers = ["Response Code", "Count", "% of Total Errors", "Severity", "Recommended Action"]

    for col, header in enumerate(error_headers, start=1):
        cell = ws_errors.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border

    for i, error in enumerate(metrics.get('error_summary', []), start=2):
        code = error['response_code']
        count = error['count']
        pct = error['pct'] * 100

        # Determine severity
        if code in ['500', '503', '504']:
            severity = "CRITICAL"
            severity_color = "FFC7CE"
        elif code in ['502', '408', '429']:
            severity = "HIGH"
            severity_color = "FFEB9C"
        else:
            severity = "MEDIUM"
            severity_color = "FFD966"

        # Recommended actions
        actions = {
            '500': "Check application logs, review recent deployments",
            '503': "Scale services, check load balancer health",
            '504': "Increase timeout settings, optimize slow queries",
            '502': "Check upstream services, network connectivity",
            '408': "Optimize long-running operations, increase timeouts",
            '429': "Implement rate limiting, optimize resource usage"
        }

        row_data = [
            code,
            count,
            f"{pct:.2f}%",
            severity,
            actions.get(code, "Review error details")
        ]

        for col, value in enumerate(row_data, start=1):
            cell = ws_errors.cell(row=i, column=col, value=value)
            cell.border = border

            if col == 4:  # Severity column
                cell.fill = PatternFill(start_color=severity_color, end_color=severity_color, fill_type="solid")
                cell.font = Font(bold=True)

    # Set column widths for errors sheet
    ws_errors.column_dimensions['A'].width = 15
    ws_errors.column_dimensions['B'].width = 10
    ws_errors.column_dimensions['C'].width = 15
    ws_errors.column_dimensions['D'].width = 12
    ws_errors.column_dimensions['E'].width = 40

    # 6. RECOMMENDATIONS TAB
    ws_recommendations = wb.create_sheet(title="Recommendations")

    # Generate recommendations
    recommendations = generate_priority_recommendations(
        metrics['global'], sla_result, metrics
    )

    ws_recommendations.merge_cells('A1:E1')
    ws_recommendations['A1'] = "PRIORITIZED RECOMMENDATIONS & ACTION PLAN"
    ws_recommendations['A1'].font = Font(size=14, bold=True, color="366092")
    ws_recommendations['A1'].alignment = Alignment(horizontal='center')

    row_num = 3
    for priority, items in recommendations.items():
        if not items:
            continue

        ws_recommendations.cell(row=row_num, column=1, value=f"{priority.upper()} PRIORITY")
        ws_recommendations.cell(row=row_num, column=1).font = Font(bold=True, size=12)

        if priority == "critical":
            fill_color = "FFC7CE"
        elif priority == "high":
            fill_color = "FFEB9C"
        elif priority == "medium":
            fill_color = "FFD966"
        else:
            fill_color = "C6EFCE"

        ws_recommendations.cell(row=row_num, column=1).fill = PatternFill(
            start_color=fill_color, end_color=fill_color, fill_type="solid"
        )

        row_num += 1

        for item in items:
            ws_recommendations.cell(row=row_num, column=1, value="•")
            ws_recommendations.cell(row=row_num, column=2, value=item['title'])
            ws_recommendations.cell(row=row_num, column=2).font = Font(bold=True)

            ws_recommendations.cell(row=row_num + 1, column=2, value=item['description'])
            ws_recommendations.cell(row=row_num + 1, column=2).font = Font(italic=True)

            # Action items
            for j, action in enumerate(item.get('action_items', []), start=1):
                ws_recommendations.cell(row=row_num + j + 1, column=3, value=f"→ {action}")

            row_num += len(item.get('action_items', [])) + 3

        row_num += 2

    # Adjust column widths
    ws_recommendations.column_dimensions['A'].width = 5
    ws_recommendations.column_dimensions['B'].width = 40
    ws_recommendations.column_dimensions['C'].width = 60

    # Save the workbook
    wb.save(output_path)
    return output_path


# =========================
# INFRASTRUCTURE METRICS (CPU/MEM) - OPTIONAL
# =========================

def infer_infra_section(file_name: str) -> str:
    """Map a JSON file name to a logical section (Load Test vs Soak Test)."""
    name = file_name.lower()
    if "loadtest" in name or "load" in name:
        return "Load Test"
    if "soaktest" in name or "soak" in name:
        return "Soak Test"
    return "Other"


def infer_infra_metric(file_name: str) -> str:
    """Infer whether this JSON contains CPU or Memory data based on its name."""
    name = file_name.lower()
    if "cpu" in name:
        return "CPU"
    if "memory" in name or "mem" in name:
        return "Memory"
    return "Metric"


def extract_infra_series(data: Any) -> Optional[Dict[str, Any]]:
    """
    Extract a multi-series time-series from infrastructure JSON.
    """
    # -------- Pattern 1: demo format with "points" ----------
    points = data.get("points")
    if isinstance(points, list):
        labels: List[str] = []
        values: List[float] = []
        for pt in points:
            if not isinstance(pt, dict):
                continue
            ts = pt.get("timestamp")
            val = pt.get("value")
            if ts is None or val is None:
                continue

            if isinstance(ts, (int, float)):
                try:
                    dt = datetime.fromtimestamp(ts)
                    label = dt.strftime("%Y-%m-d %H:%M:%S")
                except Exception:
                    label = str(ts)
            else:
                try:
                    dt = datetime.fromisoformat(str(ts))
                    label = dt.strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    label = str(ts)

            try:
                v = float(val)
            except (TypeError, ValueError):
                continue

            labels.append(label)
            values.append(v)

        if not labels:
            return None

        if values and max(values) <= 1.0:
            values = [v * 100.0 for v in values]

        return {
            "labels": labels,
            "series": [
                {"name": "Metric", "values": values}
            ],
        }

    # -------- Pattern 2: Sysdig / PromQL matrix: data.result[].values ----------
    data_block = data.get("data")
    if not isinstance(data_block, dict):
        return None

    result_list = data_block.get("result")
    if not isinstance(result_list, list) or not result_list:
        return None

    all_ts_set: Dict[float, None] = {}
    per_container_ts_values: Dict[str, Dict[float, float]] = {}

    for series in result_list:
        if not isinstance(series, dict):
            continue
        metric = series.get("metric") or {}
        container_name = (
                metric.get("kube_pod_name")
                or metric.get("container")
                or metric.get("container_name")
                or "series"
        )

        vals = series.get("values")
        if not isinstance(vals, list):
            continue

        if container_name not in per_container_ts_values:
            per_container_ts_values[container_name] = {}

        ts_map = per_container_ts_values[container_name]

        for row in vals:
            ts = None
            val = None
            if isinstance(row, (list, tuple)) and len(row) >= 2:
                ts, val = row[0], row[1]
            elif isinstance(row, dict):
                ts = row.get("timestamp")
                val = row.get("value")
            else:
                continue

            if ts is None or val is None:
                continue

            try:
                ts_float = float(ts)
                v = float(val)
            except (TypeError, ValueError):
                continue

            ts_map[ts_float] = v
            all_ts_set[ts_float] = None

    if not all_ts_set:
        return None

    all_ts_sorted = sorted(all_ts_set.keys())
    labels: List[str] = []
    for ts_float in all_ts_sorted:
        try:
            dt = datetime.fromtimestamp(ts_float)
            label = dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            label = str(ts_float)
        labels.append(label)

    series_list: List[Dict[str, Any]] = []
    global_max = 0.0

    for container_name, ts_map in per_container_ts_values.items():
        values: List[Optional[float]] = []
        for ts_float in all_ts_sorted:
            if ts_float in ts_map:
                v = ts_map[ts_float]
                values.append(v)
                if v > global_max:
                    global_max = v
            else:
                values.append(None)
        series_list.append({"name": container_name, "values": values})

    if not series_list:
        return None

    if global_max > 0.0 and global_max <= 1.0:
        for s in series_list:
            s["values"] = [v * 100.0 if v is not None else None for v in s["values"]]

    return {"labels": labels, "series": series_list}


def build_infra_metrics(file_paths: List[str]) -> Dict[str, Any]:
    """
    Build an in-memory structure for infrastructure charts from a set of JSON files.
    """
    infra: Dict[str, Dict[str, Dict[str, Any]]] = {}

    for path in file_paths:
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            continue

        file_name = os.path.basename(path)
        section = infer_infra_section(file_name)
        metric = infer_infra_metric(file_name)

        series = extract_infra_series(data)
        if not series:
            continue

        if section not in infra:
            infra[section] = {}
        infra[section][metric] = series

    return infra


def build_enhanced_infra_insights(infra_data) -> str:
    """
    Enhanced infrastructure insights with resource analysis.
    """
    if not infra_data or not isinstance(infra_data, dict):
        return """
        <div style='padding: 12px; background-color: #f3f4f6; border-radius: 8px;'>
            <div style='font-weight: 600; color: #4b5563;'>No Infrastructure Metrics</div>
            <div style='font-size: 12px; color: #6b7280; margin-top: 4px;'>Provide infrastructure JSON files with <code style='background: #e5e7eb; padding: 2px 4px; border-radius: 3px;'>--infra-json</code> for resource analysis.</div>
        </div>
        """

    load = infra_data.get("Load Test") or {}
    soak = infra_data.get("Soak Test") or {}

    def has_metric(section, key: str) -> bool:
        if not isinstance(section, dict):
            return False
        metric = section.get(key)
        return bool(
            metric
            and isinstance(metric, dict)
            and metric.get("labels")
            and metric.get("series")
        )

    active_section = None
    test_label = ""
    if has_metric(load, "CPU") or has_metric(load, "Memory"):
        active_section = load
        test_label = "Load Test"
    elif has_metric(soak, "CPU") or has_metric(soak, "Memory"):
        active_section = soak
        test_label = "Soak Test"
    else:
        return """
        <div style='padding: 12px; background-color: #f3f4f6; border-radius: 8px;'>
            <div style='font-weight: 600; color: #4b5563;'>No Usable Infrastructure Data</div>
            <div style='font-size: 12px; color: #6b7280; margin-top: 4px;'>Infrastructure JSON files were parsed, but no usable CPU/Memory series were found.</div>
        </div>
        """

    items: List[str] = []
    items.append(
        f"<div style='font-weight: 600; color: #111827; margin-bottom: 8px;'>🖥️ {test_label} Resource Analysis</div>")

    # Helper function to analyze metrics
    def analyze_metric(metric_name: str, threshold_warn: float, threshold_critical: float):
        metric_data = active_section.get(metric_name)
        if not metric_data:
            return None

        all_values = []
        container_max = {}

        for series in metric_data.get("series", []):
            name = series.get("name", "Unknown")
            values = [v for v in series.get("values", []) if v is not None]
            if values:
                container_max[name] = max(values)
                all_values.extend(values)

        if not all_values:
            return None

        avg_value = statistics.mean(all_values)
        max_value = max(all_values)
        peak_container = max(container_max.items(), key=lambda x: x[1]) if container_max else ("None", 0)

        return {
            "avg": avg_value,
            "max": max_value,
            "peak_container": peak_container,
            "status": "critical" if max_value > threshold_critical else "warning" if max_value > threshold_warn else "good"
        }

    # CPU Analysis
    cpu_analysis = analyze_metric("CPU", 80, 90)
    if cpu_analysis:
        status_color = "#dc2626" if cpu_analysis["status"] == "critical" else "#ea580c" if cpu_analysis[
                                                                                               "status"] == "warning" else "#059669"

        items.append(f"""
        <div style="margin-bottom: 12px;">
            <div style="display: flex; align-items: center; gap: 8px; margin-bottom: 4px;">
                <span style="font-size: 20px;">💻</span>
                <div style="font-weight: 500;">CPU Utilization</div>
                <div style="font-size: 11px; background: {status_color}20; color: {status_color}; padding: 2px 8px; border-radius: 10px;">{cpu_analysis['status'].upper()}</div>
            </div>
            <div style="font-size: 12px; color: #4b5563;">
                Average: <strong>{cpu_analysis['avg']:.1f}%</strong> • Peak: <strong>{cpu_analysis['max']:.1f}%</strong><br>
                Highest container: <strong>{cpu_analysis['peak_container'][0]}</strong> ({cpu_analysis['peak_container'][1]:.1f}%)
            </div>
        </div>
        """)

    # Memory Analysis
    mem_analysis = analyze_metric("Memory", 85, 95)
    if mem_analysis:
        status_color = "#dc2626" if mem_analysis["status"] == "critical" else "#ea580c" if mem_analysis[
                                                                                               "status"] == "warning" else "#059669"

        items.append(f"""
        <div style="margin-bottom: 12px;">
            <div style="display: flex; align-items: center; gap: 8px; margin-bottom: 4px;">
                <span style="font-size: 20px;">🧠</span>
                <div style="font-weight: 500;">Memory Utilization</div>
                <div style="font-size: 11px; background: {status_color}20; color: {status_color}; padding: 2px 8px; border-radius: 10px;">{mem_analysis['status'].upper()}</div>
            </div>
            <div style="font-size: 12px; color: #4b5563;">
                Average: <strong>{mem_analysis['avg']:.1f}%</strong> • Peak: <strong>{mem_analysis['max']:.1f}%</strong><br>
                Highest container: <strong>{mem_analysis['peak_container'][0]}</strong> ({mem_analysis['peak_container'][1]:.1f}%)
            </div>
        </div>
        """)

    # Resource recommendations
    if cpu_analysis and mem_analysis:
        recommendations = []

        if cpu_analysis["status"] == "critical":
            recommendations.append("Immediate CPU scaling required")
        elif cpu_analysis["status"] == "warning":
            recommendations.append("Monitor CPU usage, consider scaling")

        if mem_analysis["status"] == "critical":
            recommendations.append("Increase memory allocation")
        elif mem_analysis["status"] == "warning":
            recommendations.append("Optimize memory usage")

        if cpu_analysis["max"] < 50 and mem_analysis["max"] < 60:
            recommendations.append("Resources underutilized, could consolidate")

        if recommendations:
            items.append("<div style='font-weight: 500; margin: 12px 0 6px 0;'>🎯 Resource Recommendations</div>")
            for rec in recommendations:
                items.append(f"<div style='font-size: 12px; padding: 4px 0;'>• {rec}</div>")

    # Check for correlation with performance issues
    items.append("<div style='font-weight: 500; margin: 12px 0 6px 0;'>🔗 Performance Correlation</div>")
    if cpu_analysis and cpu_analysis["max"] > 85:
        items.append(
            "<div style='font-size: 12px; color: #dc2626;'>High CPU utilization may explain latency spikes during peak load.</div>")
    elif cpu_analysis and cpu_analysis["max"] < 40:
        items.append(
            "<div style='font-size: 12px; color: #059669;'>CPU headroom available - latency issues likely not CPU-bound.</div>")

    return "<div style='padding: 8px;'>" + "\n".join(items) + "</div>"

# =========================
# ADVANCED ANALYSIS FUNCTIONS
# =========================

def detect_anomalies_zscore(values, threshold=3.0):
    """Detect anomalies using Z-score method."""
    if len(values) < 10:
        return []

    mean = np.mean(values)
    std = np.std(values)
    if std == 0:
        return []

    z_scores = np.abs((values - mean) / std)
    anomalies = np.where(z_scores > threshold)[0].tolist()
    return anomalies


def detect_anomalies_iqr(values, threshold=1.5):
    """Detect anomalies using Interquartile Range (IQR) method."""
    if len(values) < 10:
        return []

    q1 = np.percentile(values, 25)
    q3 = np.percentile(values, 75)
    iqr = q3 - q1
    if iqr == 0:
        return []

    lower_bound = q1 - threshold * iqr
    upper_bound = q3 + threshold * iqr
    anomalies = np.where((values < lower_bound) | (values > upper_bound))[0].tolist()
    return anomalies


def calculate_trend_line(values):
    """Calculate linear trend line using least squares."""
    if len(values) < 5:
        return None, None, 0

    x = np.arange(len(values))
    slope, intercept, r_value, p_value, std_err = stats.linregress(x, values)
    return slope, intercept, r_value ** 2


def detect_periodic_pattern(values, min_period=2, max_period=20):
    """Detect periodic patterns using autocorrelation."""
    if len(values) < 30:
        return None

    try:
        autocorr = np.correlate(values - np.mean(values),
                                values - np.mean(values), mode='full')
        autocorr = autocorr[len(autocorr) // 2:]

        # Find peaks in autocorrelation
        peaks, _ = find_peaks(autocorr[:max_period], height=0.5 * np.max(autocorr))
        if len(peaks) > 0:
            period = peaks[0] + min_period
            if period < len(values) // 2:
                return period
    except:
        pass
    return None


def analyze_distribution(values, metric_name):
    """Analyze statistical distribution of values."""
    if len(values) < 10:
        return {"analysis": "Insufficient data"}

    mean_val = float(np.mean(values))
    std_val = float(np.std(values))
    skew_val = float(stats.skew(values)) if len(values) > 2 else 0

    stats_dict = {
        "metric": metric_name,
        "count": len(values),
        "mean": mean_val,
        "std": std_val,
        "min": float(np.min(values)),
        "max": float(np.max(values)),
        "cv": std_val / mean_val if mean_val > 0 else 0,
        "skewness": skew_val,
        "kurtosis": float(stats.kurtosis(values)) if len(values) > 3 else 0,
        "percentiles": {
            "p50": float(np.percentile(values, 50)),
            "p90": float(np.percentile(values, 90)),
            "p95": float(np.percentile(values, 95)),
            "p99": float(np.percentile(values, 99))
        }
    }

    # Distribution type analysis
    if skew_val > 1:
        stats_dict["distribution"] = "Right-skewed (long tail on high side)"
    elif skew_val < -1:
        stats_dict["distribution"] = "Left-skewed (long tail on low side)"
    else:
        stats_dict["distribution"] = "Approximately symmetric"

    # Variability analysis
    cv = stats_dict["cv"]
    if cv < 0.1:
        stats_dict["variability"] = "Low variability (consistent)"
    elif cv < 0.3:
        stats_dict["variability"] = "Moderate variability"
    else:
        stats_dict["variability"] = "High variability (inconsistent)"

    return stats_dict


def calculate_performance_score(metrics_dict):
    """Calculate overall performance score (0-100)."""
    score = 100

    # Error rate penalty (max -40 points)
    error_rate = metrics_dict.get("error_rate", 0) * 100
    if error_rate > 10:
        score -= 40
    elif error_rate > 5:
        score -= 30
    elif error_rate > 2:
        score -= 20
    elif error_rate > 1:
        score -= 10

    # P95 latency penalty (max -30 points)
    p95_latency = metrics_dict.get("p95_ms", 0)
    if p95_latency > 5000:
        score -= 30
    elif p95_latency > 3000:
        score -= 20
    elif p95_latency > 1000:
        score -= 10
    elif p95_latency > 500:
        score -= 5

    # Throughput consistency penalty (max -20 points)
    throughput_cv = metrics_dict.get("throughput_cv", 0)
    if throughput_cv > 0.5:
        score -= 20
    elif throughput_cv > 0.3:
        score -= 10
    elif throughput_cv > 0.2:
        score -= 5

    # Response time stability penalty (max -10 points)
    rt_stability = metrics_dict.get("rt_stability", 1)
    if rt_stability > 1.5:
        score -= 10
    elif rt_stability > 1.3:
        score -= 5

    return max(0, min(100, score))


def get_error_severity(error_code):
    """Determine severity level for error codes."""
    if error_code.startswith('5'):
        return "CRITICAL", "#fee2e2", "#dc2626"
    elif error_code in ['408', '429', '504']:
        return "HIGH", "#fef3c7", "#92400e"
    else:
        return "MEDIUM", "#f0f9ff", "#0369a1"


def get_error_cause(error_code):
    """Get likely cause for error codes."""
    error_causes = {
        '500': 'Application server error',
        '503': 'Service unavailable',
        '504': 'Gateway timeout',
        '502': 'Bad gateway',
        '408': 'Request timeout',
        '429': 'Rate limit exceeded',
        '404': 'Resource not found',
        '400': 'Bad request'
    }
    return error_causes.get(error_code, 'Unknown error')


# =========================
# ENHANCED OVERALL INSIGHTS
# =========================

def build_enhanced_overall_insights(global_stats: Dict[str, Any],
                                    sla_result: Dict[str, Any],
                                    metrics: Dict[str, Any],
                                    baseline_comparison: Optional[Dict[str, Any]] = None,
                                    infra_metrics: Optional[Dict[str, Any]] = None) -> str:
    """
    Concise overall insights with written observations.
    """
    items: List[str] = []

    # Extract key metrics
    error_rate = global_stats["error_rate"] * 100
    avg_rt = global_stats["avg_ms"]
    p95_rt = global_stats["p95_ms"]
    peak_rps = metrics.get("peak_rps", 0)

    # Generate written observations
    observations = []

    # 1. Error Rate Observation
    if error_rate > 5:
        observations.append(
            f"Error rate is critically high at {error_rate:.1f}% with {global_stats['errors']:,} errors. System stability is compromised.")
    elif error_rate > 2:
        observations.append(
            f"Error rate is elevated at {error_rate:.1f}%. Requires investigation into failure patterns.")
    else:
        observations.append(f"Error rate is acceptable at {error_rate:.1f}%. System reliability is good.")

    # 2. Response Time Observation
    if p95_rt > 3000:
        observations.append(
            f"95th percentile response time is {p95_rt:.0f}ms, exceeding the 3-second threshold. User experience is poor.")
    elif p95_rt > 1000:
        observations.append(f"95th percentile response time is {p95_rt:.0f}ms. Some users experience slow performance.")
    else:
        observations.append(f"95th percentile response time is {p95_rt:.0f}ms. Performance meets user expectations.")

    # 3. Performance Spread Observation
    if avg_rt > 0:
        spread_ratio = p95_rt / avg_rt
        if spread_ratio > 3:
            observations.append(
                f"High variance in response times (P95 is {spread_ratio:.1f}x average). Inconsistent performance.")
        elif spread_ratio > 2:
            observations.append(
                f"Moderate variance in response times (P95 is {spread_ratio:.1f}x average). Some requests are significantly slower.")
        else:
            observations.append("Response times are consistent across requests. Predictable performance.")

    # 4. Throughput Observation
    if peak_rps > 0:
        if peak_rps > 100:
            observations.append(f"System handles high load well with {peak_rps:.0f} requests per second at peak.")
        elif peak_rps > 50:
            observations.append(f"System handles moderate load with {peak_rps:.0f} requests per second at peak.")
        else:
            observations.append(f"Peak throughput is {peak_rps:.0f} RPS. Consider capacity planning for higher loads.")

    # 5. SLA Observation
    sla_passed = sla_result["passed_checks"]
    sla_total = sla_result["total_checks"]
    if sla_total > 0:
        if sla_result["overall_status"]:
            observations.append(f"All {sla_total} SLA checks passed. System meets performance requirements.")
        else:
            observations.append(f"{sla_passed}/{sla_total} SLA checks passed. Performance requirements not fully met.")

    # 6. Overall Assessment
    critical_count = sum([
        1 if error_rate > 5 else 0,
        1 if p95_rt > 3000 else 0,
        0 if sla_result["overall_status"] else 1 if sla_total > 0 else 0
    ])

    if critical_count >= 2:
        observations.append("OVERALL: Critical performance issues detected. Immediate action required.")
    elif critical_count == 1:
        observations.append("OVERALL: Performance issues present. Address before production deployment.")
    else:
        observations.append("OVERALL: Performance is satisfactory. System is ready for expected load.")

    # Limit to 5-6 key observations
    key_observations = observations[:6]

    items.append(f"""
    <div class="insight-section" style="margin-bottom: 20px;">
        <div style="display: flex; align-items: center; gap: 10px; margin-bottom: 12px;">
            <span style="font-size: 24px;">📋</span>
            <h3 style="margin: 0; color: #1e40af;">Performance Observations</h3>
        </div>

        <div style="display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 12px; margin-bottom: 15px;">
            <div style="padding: 12px; background: #f8fafc; border-radius: 8px; border-left: 4px solid {"#dc2626" if error_rate > 5 else "#d97706" if error_rate > 2 else "#059669"};">
                <div style="font-size: 12px; color: #4b5563; margin-bottom: 4px;">Error Rate</div>
                <div style="font-size: 18px; font-weight: 700; color: #111827;">{error_rate:.1f}%</div>
                <div style="font-size: 11px; color: #6b7280;">{global_stats['errors']:,} errors of {global_stats['total_requests']:,}</div>
            </div>

            <div style="padding: 12px; background: #f8fafc; border-radius: 8px; border-left: 4px solid {"#dc2626" if p95_rt > 3000 else "#d97706" if p95_rt > 1000 else "#059669"};">
                <div style="font-size: 12px; color: #4b5563; margin-bottom: 4px;">P95 Response Time</div>
                <div style="font-size: 18px; font-weight: 700; color: #111827;">{p95_rt:.0f} ms</div>
                <div style="font-size: 11px; color: #6b7280;">Avg: {avg_rt:.0f} ms</div>
            </div>

            <div style="padding: 12px; background: #f8fafc; border-radius: 8px; border-left: 4px solid {"#059669" if peak_rps > 50 else "#d97706" if peak_rps > 10 else "#dc2626"};">
                <div style="font-size: 12px; color: #4b5563; margin-bottom: 4px;">Peak Throughput</div>
                <div style="font-size: 18px; font-weight: 700; color: #111827;">{peak_rps:.0f} RPS</div>
                <div style="font-size: 11px; color: #6b7280;">Peak load capacity</div>
            </div>
        </div>

        <div style="background: #f8fafc; padding: 16px; border-radius: 8px; border: 1px solid #e5e7eb;">
            <div style="font-size: 14px; font-weight: 600; color: #1e40af; margin-bottom: 12px;">📝 Key Observations:</div>
            <div style="font-size: 13px; color: #374151; line-height: 1.6;">
                {''.join([f'<div style="margin-bottom: 8px; padding-left: 8px; border-left: 2px solid #3b82f6;">{obs}</div>' for obs in key_observations])}
            </div>

            <div style="margin-top: 16px; padding-top: 12px; border-top: 1px solid #e5e7eb;">
                <div style="font-size: 12px; font-weight: 600; color: #4b5563; margin-bottom: 6px;">📊 Test Summary:</div>
                <div style="font-size: 11px; color: #6b7280;">
                    • Total Requests: {global_stats["total_requests"]:,}<br>
                    • Test Duration: {format_duration(global_stats["duration_sec"])}<br>
                    • SLA Status: {"PASS" if sla_result["overall_status"] else "FAIL"} ({sla_passed}/{sla_total} checks)
                </div>
            </div>
        </div>
    </div>
    """)

    return "\n".join(items)

# =========================
# HTML GENERATION
# =========================

def build_html(metrics: Dict[str, Any],
                   sla_result: Dict[str, Any],
                   sla_config: Dict[str, Dict[str, float]],
                   test_name: str,
                   environment: str,
                   apdex_threshold_s: float,
                   baseline_comparison: Optional[Dict[str, Any]] = None,
                   infra_metrics: Optional[Dict[str, Any]] = None,
                   container_comparison: Optional[Dict[str, Any]] = None) -> str:
    global_stats = metrics["global"]
    per_label = metrics["per_label"]

    # --- Filter per_label to only include detected transaction names (strict transaction-only) ---
    txn_names = metrics.get("transaction_names", []) or []
    txn_per_label = [x for x in per_label if x.get("label") in txn_names]

    # If no transactions were detected, use all labels for transaction SLA view
    if not txn_per_label:
        txn_per_label = per_label

    timeseries = metrics["timeseries"]
    error_summary = metrics["error_summary"]
    error_details = metrics.get("error_details", [])

    start_dt = datetime.fromtimestamp(metrics["start_time"] / 1000.0)
    end_dt = datetime.fromtimestamp(metrics["end_time"] / 1000.0)
    generated_at = datetime.now().isoformat(timespec="seconds")

    duration_sec_val = global_stats["duration_sec"]
    duration_str = format_duration(duration_sec_val)

    ts_epochs = [ts["sec"] for ts in timeseries]
    ts_labels = [datetime.fromtimestamp(ts["sec"]).strftime("%H:%M:%S") for ts in timeseries]
    ts_rps = [ts["rps"] for ts in timeseries]
    ts_avg_ms = [ts["avg_ms"] for ts in timeseries]
    ts_active = [ts["active_threads"] if ts["active_threads"] is not None else None for ts in timeseries]
    ts_errors = [ts.get("errors", 0) for ts in timeseries]

    chart_data = {
        "ts_epoch": ts_epochs,
        "ts_labels": ts_labels,
        "ts_rps": ts_rps,
        "ts_avg_ms": ts_avg_ms,
        "ts_active": ts_active,
        "ts_errors": ts_errors,
    }

    chart_data_json = json.dumps(chart_data)
    per_label_json = json.dumps(per_label)
    baseline_json = json.dumps(baseline_comparison) if baseline_comparison else "null"
    infra_data_json = json.dumps(infra_metrics) if infra_metrics else "null"

    # Build enhanced insights
    overall_insights_html = build_enhanced_overall_insights(
        global_stats, sla_result, metrics, baseline_comparison, infra_metrics
    )
    infra_insights_html = build_enhanced_infra_insights(infra_metrics)

    # Add container comparison to infrastructure tab
    if container_comparison:
        container_comparison_html = build_container_comparison_html(container_comparison)
        infra_insights_html += container_comparison_html

    # Transaction SLA View rows based on SLA config
    transaction_sla_rows_html = []
    for ls in txn_per_label:
        sla = get_label_sla(ls["label"], sla_config)
        if not sla:
            continue

        avg = ls.get("avg")
        p90 = ls.get("p90")
        p95 = ls.get("p95")
        err_rate = ls.get("error_rate")
        count = ls.get("count", 0)
        tps = (count / duration_sec_val) if duration_sec_val > 0 else 0.0
        err_pct = (err_rate * 100.0) if err_rate is not None else 0.0

        failed = False
        if "p95_ms" in sla and p95 is not None and p95 > sla["p95_ms"]:
            failed = True
        if "avg_ms" in sla and avg is not None and avg > sla["avg_ms"]:
            failed = True
        if "error_rate" in sla and err_rate is not None and err_rate > sla["error_rate"]:
            failed = True
        if "tps_min" in sla and tps < sla["tps_min"]:
            failed = True

        status_text = "FAIL" if failed else "PASS"
        status_class = "status-fail" if failed else "status-ok"

        transaction_sla_rows_html.append(f"""
        <tr>
          <td>{ls["label"]}</td>
          <td>{format_ms(avg)}</td>
          <td>{format_ms(p90)}</td>
          <td>{format_ms(p95)}</td>
          <td>{tps:.2f}</td>
          <td>{err_pct:.2f}%</td>
          <td><span class="status-badge {status_class}">{status_text}</span></td>
        </tr>
        """)
    if transaction_sla_rows_html:
        transaction_sla_rows_html_str = "\n".join(transaction_sla_rows_html)
    else:
        transaction_sla_rows_html_str = (
            '<tr><td colspan="7" class="small">'
            'No transactions matched the SLA configuration. Update your SLA JSON to enable this view.'
            '</td></tr>'
        )

    # HTML template remains largely the same, with updated infrastructure section
    html_template = """<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>{test_name} - JMeter Performance Report </title>
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>
:root {{
    --bg:#f5f7fb;
    --card:#ffffff;
    --accent:#2563eb;
    --muted:#6b7280;
    --danger:#ef4444;
    --ok:#16a34a;
    --warning:#ea580c;
    --border:#e5e7eb;
}}
* {{
    box-sizing:border-box;
}}
body {{
    background:var(--bg);
    color:#0f172a;
    font-family:Inter,system-ui,-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Arial,sans-serif;
    margin:0;
    padding:18px;
}}
.container {{
    max-width:1200px;
    margin:0 auto;
}}
.header {{
    background:var(--card);
    padding:18px;
    border-radius:10px;
    box-shadow:0 4px 18px rgba(2,6,23,0.06);
    margin-bottom:12px;
}}
.header h1 {{
    margin:0;
    color:var(--accent);
    font-size:22px;
}}
.header .meta {{
    color:var(--muted);
    font-size:13px;
    margin-top:6px;
}}
.header .meta strong {{
    color:#111827;
}}
.small {{
    font-size:13px;
    color:var(--muted);
}}
.mono {{
    font-family:ui-monospace,SFMono-Regular,Menlo,Monaco,monospace;
    background:#f3f4f6;
    padding:2px 5px;
    border-radius:4px;
}}
.tabs {{
    display:flex;
    gap:0;
    border-radius:10px;
    overflow:hidden;
    margin:8px 0 0 0;
    border:1px solid #e5e7eb;
}}
.tab-button {{
    flex:1;
    padding:10px 14px;
    background:#f8fafc;
    border:none;
    cursor:pointer;
    text-align:center;
    font-weight:600;
    color:#334155;
    font-size:13px;
    border-right:1px solid #e5e7eb;
}}
.tab-button:last-child {{
    border-right:none;
}}
.tab-button.active {{
    background:var(--accent);
    color:#ffffff;
}}
.tab-content {{
    display:none;
    background:var(--card);
    padding:16px;
    border-radius:0 0 10px 10px;
    box-shadow:0 2px 10px rgba(2,6,23,0.04);
    margin-bottom:12px;
}}
.tab-content.active {{
    display:block;
}}
.section-title {{
    font-size:16px;
    font-weight:600;
    margin:4px 0 4px 0;
    color:#0f172a;
}}
.section-sub {{
    font-size:12px;
    color:var(--muted);
    margin:0 0 8px 0;
}}
.pill {{
    display:inline-block;
    padding:2px 8px;
    font-size:11px;
    border-radius:999px;
    background:#eef2ff;
    color:#3730a3;
    margin-left:6px;
}}
.summary-grid {{
    display:grid;
    grid-template-columns:1fr 1fr;
    gap:16px;
    margin-top:6px;
}}
.summary-card {{
    background:#f8fafc;
    padding:14px;
    border-radius:10px;
    border-left:4px solid var(--accent);
}}
.metrics-grid {{
    display:grid;
    grid-template-columns:repeat(auto-fit,minmax(180px,1fr));
    gap:10px;
    margin-top:10px;
}}
.metric-card {{
    background:linear-gradient(135deg,#667eea,#764ba2);
    color:#fff;
    padding:12px;
    border-radius:10px;
    text-align:center;
}}
.metric-value {{
    font-size:20px;
    font-weight:700;
}}
.metric-label {{
    font-size:12px;
    opacity:0.9;
    margin-top:4px;
}}
.card-plain {{
    background:var(--card);
    border-radius:10px;
    padding:10px;
    border:1px solid var(--border);
    margin-top:10px;
}}
.insights-card {{
    background:#fefce8;
    border-color:#facc15;
}}
.insights-card ul {{
    padding-left:20px;
    margin:6px 0 0 0;
    font-size:12px;
    color:#4b5563;
}}
.insights-card li {{
    margin-bottom:4px;
}}
table {{
    width:100%;
    border-collapse:collapse;
    background:var(--card);
    font-size:12px;
}}
th,td {{
    padding:8px 10px;
    border-bottom:1px solid #eef2f7;
    text-align:left;
}}
th {{
    background:#f8fafc;
    font-weight:700;
    color:#475569;
}}
tr:nth-child(even) td {{
    background:#f9fafb;
}}
.status-badge {{
    display:inline-flex;
    align-items:center;
    justify-content:center;
    min-width:48px;
    padding:2px 8px;
    border-radius:999px;
    font-size:10px;
    font-weight:600;
}}
.status-ok {{
    background-color:rgba(22,163,74,0.07);
    color:#15803d;
    border:1px solid rgba(22,163,74,0.6);
}}
.status-fail {{
    background-color:rgba(248,113,113,0.1);
    color:#b91c1c;
    border:1px solid rgba(248,113,113,0.7);
}}
.sla-overall-ok {{
    color:#15803d;
}}
.sla-overall-fail {{
    color:#b91c1c;
}}
.summary-item {{
    padding:8px;
    background:#f8fafc;
    border-radius:6px;
}}
.summary-label {{
    font-size:11px;
    color:#6b7280;
    margin-bottom:2px;
}}
.summary-value {{
    font-size:14px;
    font-weight:600;
}}
.summary-value.good {{
    color:#059669;
}}
.summary-value.warning {{
    color:#d97706;
}}
.summary-value.critical {{
    color:#dc2626;
}}
.status-banner {{
    padding:10px;
    border-radius:6px;
    margin:12px 0;
    display:flex;
    align-items:center;
    gap:8px;
}}
.status-banner.success {{
    background-color:#d1fae5;
    border:1px solid #86efac;
    color:#065f46;
}}
.status-banner.critical {{
    background-color:#fee2e2;
    border:1px solid #fca5a5;
    color:#b91c1c;
}}
.recommendation {{
    padding:10px;
    border-radius:6px;
    margin:8px 0;
}}
.recommendation.critical {{
    background-color:#fef2f2;
    border-left:4px solid #dc2626;
}}
.recommendation.high {{
    background-color:#fffbeb;
    border-left:4px solid #ea580c;
}}
.verdict {{
    padding:12px;
    border-radius:8px;
    margin:16px 0;
}}
.verdict.success {{
    background-color:#d1fae5;
    border:1px solid #86efac;
}}
.verdict.critical {{
    background-color:#fee2e2;
    border:1px solid #fca5a5;
}}
.verdict.warning {{
    background-color:#fef3c7;
    border:1px solid #fcd34d;
}}
.priority-header {{
    font-weight:600;
    margin:20px 0 8px 0;
}}
.capacity-card {{
    border-left:4px solid;
    background-color:rgba(255,255,255,0.5);
    padding:10px;
    margin:12px 0;
}}
.root-cause {{
    border-left:3px solid;
    padding:8px 0 8px 12px;
    margin:6px 0;
}}

/* CHART LAYOUT – 2-2-1 layout, bigger charts */
.chart-row {{
    display:grid;
    grid-template-columns:repeat(2, minmax(0,1fr));
    gap:18px;
    margin-top:12px;
}}
.chart-container {{
    background:var(--card);
    padding:14px;
    border-radius:10px;
    box-shadow:0 2px 8px rgba(2,6,23,0.04);
    height:420px;
    border:1px solid var(--border);
    position:relative;
}}
.chart-container canvas {{
    width:100% !important;
    height:100% !important;
    display:block;
}}
.chart-container.full-width {{
    grid-column:1 / -1;
}}
.error-badge {{
    background:var(--danger);
    color:#fff;
    padding:6px 10px;
    border-radius:18px;
    font-size:12px;
}}
.footer {{
    margin-top:8px;
    font-size:11px;
    color:var(--muted);
    text-align:center;
}}
.filter-row {{
    display:flex;
    flex-wrap:wrap;
    gap:10px;
    align-items:flex-end;
    margin-top:6px;
}}
.filter-group {{
    display:flex;
    flex-direction:column;
    gap:4px;
    font-size:12px;
    color:var(--muted);
}}
.filter-group input, .filter-group select {{
    padding:4px 8px;
    border-radius:6px;
    border:1px solid var(--border);
    font-size:12px;
}}
.filter-actions button {{
    padding:6px 10px;
    border-radius:6px;
    border:none;
    font-size:12px;
    cursor:pointer;
}}
.btn-primary {{
    background:var(--accent);
    color:#fff;
}}
.btn-secondary {{
    background:#e5e7eb;
    color:#374151;
}}
@media (max-width:900px) {{
    .summary-grid {{
        grid-template-columns:1fr;
    }}
    .chart-row {{
        grid-template-columns:1fr;
    }}
    .chart-container {{
        height:340px;
    }}
}}
</style>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
</head>
<body>
<div class="container">

  <div class="header">
    <h1>JMeter Performance Report </h1>
    <div class="meta">
      Generated: <span class="mono">{generated_at}</span> ·
      Samples: <strong>{total_requests}</strong> ·
      Duration: <strong>{duration_str}</strong>
    </div>
    <div class="meta">
      Test: <strong>{test_name}</strong> · Environment: <strong>{environment}</strong> ·
      Start: <span class="mono">{start_time}</span> · End: <span class="mono">{end_time}</span>
    </div>
  </div>

  <!-- GLOBAL TIME WINDOW FILTER -->
  <div class="card-plain" style="margin-top:8px;margin-bottom:8px;">
    <div class="small" style="margin-bottom:4px;">Time Window Filter (applies to charts)</div>
    <div class="filter-row">
      <div class="filter-group">
        <label for="timeWindowSelectTop">Window</label>
        <select id="timeWindowSelectTop">
          <option value="full">Full Run</option>
          <option value="first10">First 10 Minutes</option>
          <option value="last10">Last 10 Minutes</option>
          <option value="firstHalf">First Half</option>
          <option value="secondHalf">Second Half</option>
          <option value="custom">Custom Range</option>
        </select>
      </div>
      <div class="filter-group">
        <label for="timeWindowCustomStart">Custom Start</label>
        <input type="datetime-local" id="timeWindowCustomStart">
      </div>
      <div class="filter-group">
        <label for="timeWindowCustomEnd">Custom End</label>
        <input type="datetime-local" id="timeWindowCustomEnd">
      </div>
      <div class="filter-actions" style="display:flex;gap:8px;">
        <button id="applyCustomWindow" class="btn-primary">Apply Custom Range</button>
        <button id="clearTimeWindow" class="btn-secondary">Clear</button>
      </div>
    </div>
  </div>

  <!-- OVERALL INSIGHTS -->
  <div class="card-plain insights-card" style="margin-top:8px;margin-bottom:8px;">
    <div class="section-title" style="margin-top:0;">Overall Insights &amp; Recommendations</div>
    <p class="section-sub">
      Enhanced analysis with health score, root cause identification, and prioritized action items.
    </p>
    {overall_insights_html}
  </div>

  <!-- Tabs -->
  <div class="tabs" role="tablist">
    <button class="tab-button active" data-tab="dashboard">Dashboard</button>
    <button class="tab-button" data-tab="graphs">Graphs</button>
    <button class="tab-button" data-tab="errors">Error Report</button>
    <button class="tab-button" data-tab="infrastructure">Infrastructure</button>
    <button class="tab-button" data-tab="comparison">Comparison</button>
  </div>

  <!-- DASHBOARD TAB -->
  <div id="tab-dashboard" class="tab-content active">
    <div class="section-title">Dashboard <span class="pill">Overview</span></div>
    <p class="section-sub">High-level view of the test run and transaction-level metrics.</p>

    <div class="summary-grid">
      <div class="summary-card">
        <h3>Test Summary</h3>
        <div style="margin-top:8px">
          <div class="small">Total Samples: <strong>{total_requests}</strong></div>
          <div class="small">Success: <strong>{successes}</strong> · Failures: <strong>{errors}</strong></div>
          <div class="small">Error Rate: <strong>{error_rate_pct:.2f}%</strong></div>
          <div class="small">Avg Response: <strong>{avg_ms:.1f} ms</strong></div>
          <div class="small">P95 Response: <strong>{p95_ms:.1f} ms</strong></div>
        </div>
      </div>
      <div class="summary-card">
        <h3>SLA Summary</h3>
        <div style="margin-top:8px">
          <div class="{sla_overall_class}">Overall SLA: <strong>{sla_overall_text}</strong></div>
          <div class="small" style="margin-top:6px">
            Checks passed: <strong>{sla_passed}</strong> / {sla_total}
            ({sla_pct:.1f}%)
          </div>
          <div class="small" style="margin-top:6px">
            APDEX (T = {apdex_threshold:.1f}s): <strong>{apdex_value:.2f}</strong>
          </div>
        </div>
      </div>
    </div>

    <div class="metrics-grid">
      <div class="metric-card">
        <div class="metric-value">{peak_rps}</div>
        <div class="metric-label">Peak Throughput (req/s)</div>
      </div>
      <div class="metric-card">
        <div class="metric-value">{avg_ms:.1f}</div>
        <div class="metric-label">Average Response (ms)</div>
      </div>
      <div class="metric-card">
        <div class="metric-value">{p95_ms:.1f}</div>
        <div class="metric-label">95th Percentile (ms)</div>
      </div>
      <div class="metric-card">
        <div class="metric-value">{error_rate_pct:.2f}%</div>
        <div class="metric-label">Error Rate</div>
      </div>
    </div>

    <!-- TRANSACTION SLA VIEW -->
    <div class="card-plain" style="margin-top:14px;">
      <div class="section-title" style="margin-top:0;">Transaction SLA View</div>
      <p class="section-sub">
        Transactions evaluated against configured SLA values. Status is PASS/FAIL based on P95, Avg, error rate and optional TPS thresholds.
      </p>
      <div style="max-height:260px;overflow:auto;">
        <table>
          <thead>
            <tr>
              <th>Transaction</th>
              <th>Avg (ms)</th>
              <th>P90 (ms)</th>
              <th>P95 (ms)</th>
              <th>Throughput (req/s)</th>
              <th>Error %</th>
              <th>Status</th>
            </tr>
          </thead>
          <tbody>
            {transaction_sla_rows_html}
          </tbody>
        </table>
      </div>
    </div>

    <!-- TRANSACTION STATISTICS -->
    <div class="card-plain" style="margin-top:14px;">
      <div class="section-title" style="margin-top:0;">Transaction Statistics</div>
      <p class="section-sub">
        Detailed statistics per transaction, with filters by average response time, throughput and error percentage.
      </p>

      <div class="filter-row">
        <div class="filter-group">
          <label for="filterAvgMax">Max Avg Response (ms)</label>
          <input id="filterAvgMax" type="number" placeholder="e.g. 2000">
        </div>
        <div class="filter-group">
          <label for="filterTpsMin">Min TPS (Hits/sec)</label>
          <input id="filterTpsMin" type="number" step="0.01" placeholder="e.g. 5">
        </div>
        <div class="filter-group">
          <label for="filterErrMax">Max Error %</label>
          <input id="filterErrMax" type="number" step="0.01" placeholder="e.g. 2">
        </div>
        <div class="filter-actions" style="display:flex;gap:8px;">
          <button id="applySlaFilter" class="btn-primary">Apply</button>
          <button id="resetSlaFilter" class="btn-secondary">Clear</button>
        </div>
      </div>

      <div style="max-height:320px;overflow:auto;margin-top:10px;">
        <table id="perTxnTable">
          <thead>
            <tr>
              <th>Label</th>
              <th>Count</th>
              <th>Min (ms)</th>
              <th>Avg (ms)</th>
              <th>P50 (ms)</th>
              <th>P90 (ms)</th>
              <th>P95 (ms)</th>
              <th>P99 (ms)</th>
              <th>Max (ms)</th>
              <th>Throughput (req/min)</th>
              <th>Hits/sec</th>
              <th>Error Rate</th>
            </tr>
          </thead>
          <tbody>
            {per_label_rows_html}
          </tbody>
        </table>
      </div>
    </div>
  </div>

  <!-- GRAPHS TAB -->
  <div id="tab-graphs" class="tab-content">
    <div class="section-title">Time-series &amp; Trends <span class="pill">Graphs</span></div>
    <p class="section-sub">
      Throughput, response time, errors, and active threads over time.
      Use the time window filter to focus on specific portions of the test.
    </p>

    <div class="card-plain" style="margin-top:6px;margin-bottom:8px;">
      <div class="small" style="margin-bottom:4px;">Time Window Filter (applies to charts below)</div>
      <div class="filter-row">
        <div class="filter-group">
          <label for="timeWindowSelect">Window</label>
          <select id="timeWindowSelect">
            <option value="full">Full Run</option>
            <option value="first10">First 10 Minutes</option>
            <option value="last10">Last 10 Minutes</option>
            <option value="firstHalf">First Half</option>
            <option value="secondHalf">Second Half</option>
            <option value="custom">Custom Range</option>
          </select>
        </div>
      </div>
    </div>

    <!-- ROW 1 -->
    <div class="chart-row">
      <div class="chart-container">
        <div class="small" style="margin-bottom:6px;">Requests Per Second</div>
        <canvas id="rpsChart"></canvas>
      </div>
      <div class="chart-container">
        <div class="small" style="margin-bottom:6px;">Average Response Time (ms)</div>
        <canvas id="rtChart"></canvas>
      </div>
    </div>

    <!-- ROW 2 -->
    <div class="chart-row">
      <div class="chart-container">
        <div class="small" style="margin-bottom:6px;">Errors Per Second</div>
        <canvas id="errChart"></canvas>
      </div>
      <div class="chart-container">
        <div class="small" style="margin-bottom:6px;">Response Time vs Errors Per Second</div>
        <canvas id="rtErrChart"></canvas>
      </div>
    </div>

    <!-- ROW 3 -->
    <div class="chart-row">
      <div class="chart-container full-width">
        <div class="small" style="margin-bottom:6px;">Throughput vs Threads</div>
        <canvas id="rpsThreadChart"></canvas>
      </div>
    </div>
  </div>

  <!-- ERRORS TAB -->
  <div id="tab-errors" class="tab-content">
    <div class="section-title">Error Report <span class="pill">Stability</span></div>
    <p class="section-sub">Grouped by response code for this test run, plus transaction-level error details.</p>

    <div style="display:flex;gap:12px;align-items:center;margin-bottom:10px;">
      <div class="error-badge">{errors} Errors</div>
      <div class="small">Non-success responses by HTTP code and transaction.</div>
    </div>

    <!-- SUMMARY BY RESPONSE CODE -->
    <div style="max-height:260px;overflow:auto;">
      <table>
        <thead>
          <tr>
            <th>Response Code</th>
            <th>Count</th>
            <th>Percentage of Errors</th>
          </tr>
        </thead>
        <tbody>
          {error_rows_html}
        </tbody>
      </table>
    </div>

    <!-- DETAILED BREAKDOWN BY TRANSACTION & MESSAGE -->
    <div style="max-height:260px;overflow:auto;margin-top:10px;">
      <table>
        <thead>
          <tr>
            <th>Transaction</th>
            <th>Response Code</th>
            <th>Error Message</th>
            <th>Count</th>
            <th>Percentage of Errors</th>
          </tr>
        </thead>
        <tbody>
          {error_detail_rows_html}
        </tbody>
      </table>
    </div>
  </div>

  <!-- INFRASTRUCTURE TAB -->
  <div id="tab-infrastructure" class="tab-content">
    <div class="section-title">Infrastructure Metrics (Load / Soak)</div>
    <p class="section-sub">
      CPU and Memory utilization during test execution. Graphs will automatically use Load Test
      data if present, otherwise Soak Test data.
    </p>

    <!-- INFRASTRUCTURE INSIGHTS -->
    <div class="card-plain insights-card" style="margin-top:6px;">
      <div class="section-title" style="margin-top:0;">Insights &amp; Recommendations</div>
      <p class="section-sub">
        Resource utilization analysis and correlation with performance.
      </p>
      {infra_insights_html}
    </div>

    <!-- SINGLE TEST METRICS ROW (will be Load OR Soak) -->
    <div class="section-title" style="margin-top:12px;">Test Metrics</div>
    <p class="section-sub">
      Charts below will show metrics for either the Load Test or the Soak Test, depending on which
      infrastructure JSON files are provided.
    </p>

    <div class="chart-row">
      <div class="chart-container">
        <div class="small infra-chart-title" data-metric="CPU" style="margin-bottom:6px;">
          CPU Utilization
        </div>
        <canvas id="loadCpuChart"></canvas>
      </div>
      <div class="chart-container">
        <div class="small infra-chart-title" data-metric="Memory" style="margin-bottom:6px;">
          Memory Utilization
        </div>
        <canvas id="loadMemChart"></canvas>
      </div>
    </div>

    <!-- SINGLE CONTAINER LEGEND -->
    <div class="card-plain" style="margin-top:14px;" id="infraContainerLegend">
      <div class="section-title" style="margin-top:0;">Container Lines &amp; Colors</div>
      <p class="section-sub">
        Each color corresponds to one container and is reused across all CPU/Memory charts.
      </p>
      <div id="infraContainerLegendContent" class="small"></div>
    </div>
  </div>

  <!-- COMPARISON TAB -->
  <div id="tab-comparison" class="tab-content">
    <div class="section-title">Baseline Comparison <span class="pill">Trends</span></div>
    <p class="section-sub">Compare this run against a baseline CSV/JTL (if provided).</p>

    <div class="card-plain" style="margin-top:10px;max-height:260px;overflow:auto;">
      <div id="baselineSection"></div>
      <p class="small" style="margin-top:6px;">
        Provide <span class="mono">--baseline baseline.csv</span> to enable this comparison.
      </p>
    </div>
  </div>

  <div class="footer">
    Generated by JMeter HTML Performance Dashboard v3.7 · Enhanced Insights Engine · Container Comparison · Excel Export · Python · Chart.js
  </div>

</div>

<script>
document.addEventListener('DOMContentLoaded', function() {{
  const chartData = {chart_data_json};
  const perLabelData = {per_label_json};
  const baselineComparison = {baseline_json};
  const infraData = {infra_data_json};

  // --- Charts ---
  const ctxRps = document.getElementById('rpsChart').getContext('2d');
  const rpsChart = new Chart(ctxRps, {{
      type: 'line',
      data: {{
          labels: chartData.ts_labels,
          datasets: [{{
              label: 'RPS',
              data: chartData.ts_rps,
              borderColor: 'rgba(37,99,235,1)',
              backgroundColor: 'rgba(37,99,235,0.15)',
              borderWidth: 1.6,
              pointRadius: 0,
              tension: 0.25,
              fill: true
          }}]
      }},
      options: {{
          responsive: true,
          plugins: {{
              legend: {{ display: false }},
              tooltip: {{ mode: 'index', intersect: false }}
          }},
          interaction: {{ mode: 'index', intersect: false }},
          maintainAspectRatio: false,
          scales: {{
              x: {{
                  ticks: {{ maxTicksLimit: 10 }},
                  grid: {{ color: 'rgba(148,163,184,0.4)' }}
              }},
              y: {{
                  title: {{ display: true, text: 'Requests / Second' }},
                  grid: {{ color: 'rgba(148,163,184,0.4)' }}
              }}
          }}
      }}
  }});

  const ctxRt = document.getElementById('rtChart').getContext('2d');
  const rtChart = new Chart(ctxRt, {{
      type: 'line',
      data: {{
          labels: chartData.ts_labels,
          datasets: [{{
              label: 'Avg Response Time (ms)',
              data: chartData.ts_avg_ms,
              borderColor: 'rgba(79,70,229,1)',
              backgroundColor: 'rgba(79,70,229,0.16)',
              borderWidth: 1.6,
              pointRadius: 0,
              tension: 0.25,
              fill: true
          }}]
      }},
      options: {{
          responsive: true,
          plugins: {{
              legend: {{ display: false }},
              tooltip: {{ mode: 'index', intersect: false }}
          }},
          interaction: {{ mode: 'index', intersect: false }},
          maintainAspectRatio: false,
          scales: {{
              x: {{
                  ticks: {{ maxTicksLimit: 10 }},
                  grid: {{ color: 'rgba(148,163,184,0.4)' }}
              }},
              y: {{
                  title: {{ display: true, text: 'Avg Response Time (ms)' }},
                  grid: {{ color: 'rgba(148,163,184,0.4)' }}
              }}
          }}
      }}
  }});

  const ctxErr = document.getElementById('errChart').getContext('2d');
  const errChart = new Chart(ctxErr, {{
      type: 'line',
      data: {{
          labels: chartData.ts_labels,
          datasets: [{{
              label: 'Errors / Second',
              data: chartData.ts_errors,
              borderColor: 'rgba(239,68,68,1)',
              backgroundColor: 'rgba(239,68,68,0.18)',
              borderWidth: 1.6,
              pointRadius: 0,
              tension: 0.25,
              fill: true
          }}]
      }},
      options: {{
          responsive: true,
          plugins: {{
              legend: {{ display: false }},
              tooltip: {{ mode: 'index', intersect: false }}
          }},
          interaction: {{ mode: 'index', intersect: false }},
          maintainAspectRatio: false,
          scales: {{
              x: {{
                  ticks: {{ maxTicksLimit: 10 }},
                  grid: {{ color: 'rgba(148,163,184,0.4)' }}
              }},
              y: {{
                  title: {{ display: true, text: 'Errors / Second' }},
                  grid: {{ color: 'rgba(148,163,184,0.4)' }}
              }}
          }}
      }}
  }});

  const ctxRtErr = document.getElementById('rtErrChart').getContext('2d');
  const rtErrChart = new Chart(ctxRtErr, {{
      type: 'line',
      data: {{
          labels: chartData.ts_labels,
          datasets: [
              {{
                  label: 'Avg Response Time (ms)',
                  data: chartData.ts_avg_ms,
                  borderColor: 'rgba(79,70,229,1)',
                  backgroundColor: 'rgba(79,70,229,0.0)',
                  borderWidth: 1.6,
                  pointRadius: 0,
                  tension: 0.25,
                  fill: false,
                  yAxisID: 'y'
              }},
              {{
                  label: 'Errors / Second',
                  data: chartData.ts_errors,
                  borderColor: 'rgba(239,68,68,1)',
                  backgroundColor: 'rgba(239,68,68,0.0)',
                  borderWidth: 1.6,
                  pointRadius: 0,
                  tension: 0.25,
                  fill: false,
                  yAxisID: 'y1'
              }}
          ]
      }},
      options: {{
          responsive: true,
          plugins: {{
              legend: {{ display: true }},
              tooltip: {{ mode: 'index', intersect: false }}
          }},
          interaction: {{ mode: 'index', intersect: false }},
          maintainAspectRatio: false,
          scales: {{
              x: {{
                  ticks: {{ maxTicksLimit: 10 }},
                  grid: {{ color: 'rgba(148,163,184,0.4)' }}
              }},
              y: {{
                  title: {{ display: true, text: 'Response Time (ms)' }},
                  position: 'left',
                  grid: {{ color: 'rgba(148,163,184,0.4)' }}
              }},
              y1: {{
                  title: {{ display: true, text: 'Errors / Second' }},
                  position: 'right',
                  grid: {{ drawOnChartArea: false }}
              }}
          }}
      }}
  }});

  const ctxRpsThread = document.getElementById('rpsThreadChart').getContext('2d');
  const rpsThreadChart = new Chart(ctxRpsThread, {{
      type: 'line',
      data: {{
          labels: chartData.ts_labels,
          datasets: [
              {{
                  label: 'RPS',
                  data: chartData.ts_rps,
                  borderColor: 'rgba(37,99,235,1)',
                  backgroundColor: 'rgba(37,99,235,0.0)',
                  borderWidth: 1.6,
                  pointRadius: 0,
                  tension: 0.25,
                  fill: false,
                  yAxisID: 'y'
              }},
              {{
                  label: 'Active Threads',
                  data: chartData.ts_active,
                  borderColor: 'rgba(34,197,94,1)',
                  backgroundColor: 'rgba(34,197,94,0.0)',
                  borderWidth: 1.6,
                  pointRadius: 0,
                  tension: 0.25,
                  fill: false,
                  yAxisID: 'y1'
              }}
          ]
      }},
      options: {{
          responsive: true,
          plugins: {{
              legend: {{ display: true }},
              tooltip: {{ mode: 'index', intersect: false }}
          }},
          interaction: {{ mode: 'index', intersect: false }},
          maintainAspectRatio: false,
          scales: {{
              x: {{
                  ticks: {{ maxTicksLimit: 10 }},
                  grid: {{ color: 'rgba(148,163,184,0.4)' }}
              }},
              y: {{
                  title: {{ display: true, text: 'Requests / Second' }},
                  position: 'left',
                  grid: {{ color: 'rgba(148,163,184,0.4)' }}
              }},
              y1: {{
                  title: {{ display: true, text: 'Active Threads' }},
                  position: 'right',
                  grid: {{ drawOnChartArea: false }}
              }}
          }}
      }}
  }});

  /* ========================
     INFRASTRUCTURE CHARTS
     ======================== */
  const infraColorPalette = [
      'rgba(37,99,235,1)',
      'rgba(34,197,94,1)',
      'rgba(234,179,8,1)',
      'rgba(239,68,68,1)',
      'rgba(139,92,246,1)',
      'rgba(14,165,233,1)',
      'rgba(244,114,182,1)',
      'rgba(148,163,184,1)'
  ];

  const infraColorMap = new Map();

  function getInfraColor(name) {{
      const key = name || 'Unknown';
      if (infraColorMap.has(key)) return infraColorMap.get(key);
      const color = infraColorPalette[infraColorMap.size % infraColorPalette.length];
      infraColorMap.set(key, color);
      return color;
  }}

  function makeInfraChart(canvasId, series, titleText) {{
      const canvas = document.getElementById(canvasId);
      if (!canvas || !series || !series.labels || !series.series || series.labels.length === 0) {{
          if (canvas) {{
              const parent = canvas.parentElement;
              if (parent) {{
                  const msg = document.createElement('div');
                  msg.className = 'small';
                  msg.textContent = 'No data available for this metric.';
                  parent.appendChild(msg);
              }}
          }}
          return;
      }}

      const ctx = canvas.getContext('2d');

      const datasets = series.series.map((s) => {{
          const color = getInfraColor(s.name);
          return {{
              label: s.name,
              data: s.values,
              borderColor: color,
              backgroundColor: 'rgba(0,0,0,0)',
              borderWidth: 1.6,
              pointRadius: 0,
              tension: 0.25,
              spanGaps: true,
              fill: false
          }};
      }});

      new Chart(ctx, {{
          type: 'line',
          data: {{
              labels: series.labels,
              datasets: datasets
          }},
          options: {{
              responsive: true,
              plugins: {{
                  legend: {{ display: false }},
                  tooltip: {{ mode: 'index', intersect: false }}
              }},
              interaction: {{ mode: 'index', intersect: false }},
              maintainAspectRatio: false,
              scales: {{
                  x: {{
                      ticks: {{ maxTicksLimit: 8 }},
                      grid: {{ color: 'rgba(148,163,184,0.4)' }}
                  }},
                  y: {{
                      title: {{ display: true, text: 'Utilization (%)' }},
                      min: 0,
                      max: 100,
                      grid: {{ color: 'rgba(148,163,184,0.4)' }}
                  }}
              }}
          }}
      }});
  }}

  if (infraData) {{
      const load = infraData["Load Test"] || {{}};
      const soak = infraData["Soak Test"] || {{}};

      function hasMetric(section, key) {{
          return section && section[key] && section[key].labels && section[key].labels.length > 0;
      }}

      let activeSection = null;
      let labelSuffix = '';

      if (hasMetric(load, "CPU") || hasMetric(load, "Memory")) {{
          activeSection = load;
          labelSuffix = ' (Load Test)';
      }} else if (hasMetric(soak, "CPU") || hasMetric(soak, "Memory")) {{
          activeSection = soak;
          labelSuffix = ' (Soak Test)';
      }}

      document.querySelectorAll('.infra-chart-title').forEach(el => {{
          const metric = el.getAttribute('data-metric') || '';
          el.textContent = metric + ' Utilization' + labelSuffix;
      }});

      if (activeSection) {{
          if (hasMetric(activeSection, "CPU")) {{
              makeInfraChart(
                  'loadCpuChart',
                  activeSection["CPU"],
                  'CPU Utilization' + labelSuffix
              );
          }}

          if (hasMetric(activeSection, "Memory")) {{
              makeInfraChart(
                  'loadMemChart',
                  activeSection["Memory"],
                  'Memory Utilization' + labelSuffix
              );
          }}
      }}

      const legendHost = document.getElementById('infraContainerLegendContent');
      if (legendHost && infraColorMap.size > 0) {{
          let html = '<div style="display:flex;flex-wrap:wrap;gap:8px;">';
          infraColorMap.forEach((color, name) => {{
              html += ''
                  + '<div style="display:flex;align-items:center;gap:4px;'
                  + 'padding:2px 6px;border-radius:999px;border:1px solid #e5e7eb;">'
                  +   '<span style="width:10px;height:10px;border-radius:999px;background:' + color + ';"></span>'
                  +   '<span>' + name + '</span>'
                  + '</div>';
          }});
          html += '</div>';
          legendHost.innerHTML = html;
      }}
  }}

  // --- Baseline comparison ---
  if (baselineComparison && baselineComparison.rows && baselineComparison.rows.length > 0) {{
      const container = document.getElementById('baselineSection');
      let html = '<table><thead><tr>' +
                 '<th>Transaction</th>' +
                 '<th>Baseline P95 (ms)</th>' +
                 '<th>Current P95 (ms)</th>' +
                 '<th>Δ P95 (ms)</th>' +
                 '<th>Baseline Error</th>' +
                 '<th>Current Error</th>' +
                 '<th>Δ Error</th>' +
                 '</tr></thead><tbody>';

      baselineComparison.rows.forEach(row => {{
          const fmtMs = x => (x === null || x === undefined) ? '-' : x.toFixed(1);
          const fmtPct = x => (x === null || x === undefined) ? '-' : (x * 100).toFixed(2) + '%';

          let deltaClass = '';
          if (row.delta_p95 !== null && row.delta_p95 !== undefined) {{
              if (row.delta_p95 > 0) deltaClass = 'status-badge status-fail';
              else if (row.delta_p95 < 0) deltaClass = 'status-badge status-ok';
          }}

          html += '<tr>' +
              '<td>' + row.label + '</td>' +
              '<td>' + fmtMs(row.baseline_p95) + '</td>' +
              '<td>' + fmtMs(row.current_p95) + '</td>' +
              '<td><span class=\"' + deltaClass + '\">' +
                  ((row.delta_p95 === null || row.delta_p95 === undefined)
                      ? '-' : row.delta_p95.toFixed(1)) +
              '</span></td>' +
              '<td>' + fmtPct(row.baseline_error) + '</td>' +
              '<td>' + fmtPct(row.current_error) + '</td>' +
              '<td>' + fmtPct(row.delta_error) + '</td>' +
              '</tr>';
      }});
      html += '</tbody></table>';
      container.innerHTML = html;
  }} else {{
      const container = document.getElementById('baselineSection');
      if (container) {{
          container.innerHTML =
              '<p class="small">No baseline data provided.</p>';
      }}
  }}

  // --- Tabs ---
  const tabButtons = document.querySelectorAll('.tab-button');
  const tabContents = document.querySelectorAll('.tab-content');
  tabButtons.forEach(btn => {{
      btn.addEventListener('click', () => {{
          const tab = btn.dataset.tab;
          tabButtons.forEach(b => b.classList.remove('active'));
          btn.classList.add('active');
          tabContents.forEach(c => {{
              c.classList.toggle('active', c.id === 'tab-' + tab);
          }});
      }});
  }});

  // --- Transaction Statistics filter ---
  const filterAvgMax = document.getElementById('filterAvgMax');
  const filterTpsMin = document.getElementById('filterTpsMin');
  const filterErrMax = document.getElementById('filterErrMax');
  const perTxnTableBody = document.querySelector('#perTxnTable tbody');

  function applySlaFilter() {{
      if (!perTxnTableBody) return;
      const avgMax = parseFloat(filterAvgMax.value);
      const tpsMin = parseFloat(filterTpsMin.value);
      const errMax = parseFloat(filterErrMax.value);

      const rows = perTxnTableBody.querySelectorAll('tr');
      rows.forEach(row => {{
          const avg = parseFloat(row.dataset.avgMs || 'NaN');
          const tps = parseFloat(row.dataset.tps || 'NaN');
          const err = parseFloat(row.dataset.errorPct || 'NaN');

          let visible = true;

          if (!isNaN(avgMax) && !isNaN(avg) && avg > avgMax) {{
              visible = false;
          }}
          if (!isNaN(tpsMin) && (isNaN(tps) || tps < tpsMin)) {{
              visible = false;
          }}
          if (!isNaN(errMax) && !isNaN(err) && err > errMax) {{
              visible = false;
          }}

          row.style.display = visible ? '' : 'none';
      }});
  }}

  function resetSlaFilter() {{
      if (!perTxnTableBody) return;
      filterAvgMax.value = '';
      filterTpsMin.value = '';
      filterErrMax.value = '';
      perTxnTableBody.querySelectorAll('tr').forEach(row => row.style.display = '');
  }}

  const applySlaBtn = document.getElementById('applySlaFilter');
  const resetSlaBtn = document.getElementById('resetSlaFilter');
  if (applySlaBtn) {{
      applySlaBtn.addEventListener('click', (e) => {{
          e.preventDefault();
          applySlaFilter();
      }});
  }}
  if (resetSlaBtn) {{
      resetSlaBtn.addEventListener('click', (e) => {{
          e.preventDefault();
          resetSlaFilter();
      }});
  }}

  // --- Time Window Filter (presets + custom) ---
  const timeWindowSelect = document.getElementById('timeWindowSelect');
  const timeWindowSelectTop = document.getElementById('timeWindowSelectTop');
  const customStartInput = document.getElementById('timeWindowCustomStart');
  const customEndInput = document.getElementById('timeWindowCustomEnd');
  const applyCustomBtn = document.getElementById('applyCustomWindow');
  const clearTimeWindowBtn = document.getElementById('clearTimeWindow');

  function applyTimeWindow(mode, customLow = null, customHigh = null) {{
      const epochs = chartData.ts_epoch;
      if (!epochs || epochs.length === 0) return;

      const labelsFull = chartData.ts_labels;
      const rpsFull = chartData.ts_rps;
      const avgFull = chartData.ts_avg_ms;
      const errFull = chartData.ts_errors;
      const activeFull = chartData.ts_active;

      const minSec = epochs[0];
      const maxSec = epochs[epochs.length - 1];
      const span = maxSec - minSec;

      let low = minSec;
      let high = maxSec;

      if (mode === 'custom' && customLow !== null && customHigh !== null) {{
          low = Math.max(minSec, Math.min(customLow, customHigh));
          high = Math.min(maxSec, Math.max(customLow, customHigh));
      }} else {{
          if (mode === 'first10') {{
              high = Math.min(maxSec, minSec + 600);
          }} else if (mode === 'last10') {{
              low = Math.max(minSec, maxSec - 600);
          }} else if (mode === 'firstHalf') {{
              high = minSec + span / 2;
          }} else if (mode === 'secondHalf') {{
              low = minSec + span / 2;
          }} else {{
              low = minSec;
              high = maxSec;
          }}
      }}

      const newLabels = [];
      const newRps = [];
      const newAvg = [];
      const newErr = [];
      const newActive = [];

      for (let i = 0; i < epochs.length; i++) {{
          const t = epochs[i];
          if (t >= low && t <= high) {{
              newLabels.push(labelsFull[i]);
              newRps.push(rpsFull[i]);
              newAvg.push(avgFull[i]);
              newErr.push(errFull[i]);
              newActive.push(activeFull[i]);
          }}
      }}

      const L = newLabels.length > 0 ? newLabels : labelsFull;
      const R = newLabels.length > 0 ? newRps : rpsFull;
      const A = newLabels.length > 0 ? newAvg : avgFull;
      const E = newLabels.length > 0 ? newErr : errFull;
      const T = newLabels.length > 0 ? newActive : activeFull;

      rpsChart.data.labels = L;
      rpsChart.data.datasets[0].data = R;
      rpsChart.update();

      rtChart.data.labels = L;
      rtChart.data.datasets[0].data = A;
      rtChart.update();

      errChart.data.labels = L;
      errChart.data.datasets[0].data = E;
      errChart.update();

      rtErrChart.data.labels = L;
      rtErrChart.data.datasets[0].data = A;
      rtErrChart.data.datasets[1].data = E;
      rtErrChart.update();

      rpsThreadChart.data.labels = L;
      rpsThreadChart.data.datasets[0].data = R;
      rpsThreadChart.data.datasets[1].data = T;
      rpsThreadChart.update();
  }}

  function setTimeWindow(mode) {{
      if (timeWindowSelect) timeWindowSelect.value = mode;
      if (timeWindowSelectTop) timeWindowSelectTop.value = mode;
      if (mode !== 'custom') {{
          applyTimeWindow(mode);
      }}
  }}

  function applyCustomWindow() {{
      if (!customStartInput || !customEndInput) return;
      const startVal = customStartInput.value;
      const endVal = customEndInput.value;
      if (!startVal || !endVal) return;
      const startEpoch = Math.floor(new Date(startVal).getTime() / 1000);
      const endEpoch = Math.floor(new Date(endVal).getTime() / 1000);
      if (isNaN(startEpoch) || isNaN(endEpoch)) return;

      if (timeWindowSelect) timeWindowSelect.value = 'custom';
      if (timeWindowSelectTop) timeWindowSelectTop.value = 'custom';
      applyTimeWindow('custom', startEpoch, endEpoch);
  }}

  function clearTimeWindow() {{
      if (customStartInput) customStartInput.value = '';
      if (customEndInput) customEndInput.value = '';
      setTimeWindow('full');
  }}

  if (timeWindowSelect) {{
      timeWindowSelect.addEventListener('change', () => {{
          const mode = timeWindowSelect.value;
          setTimeWindow(mode);
      }});
  }}
  if (timeWindowSelectTop) {{
      timeWindowSelectTop.addEventListener('change', () => {{
          const mode = timeWindowSelectTop.value;
          setTimeWindow(mode);
      }});
  }}
  if (applyCustomBtn) {{
      applyCustomBtn.addEventListener('click', (e) => {{
          e.preventDefault();
          applyCustomWindow();
      }});
  }}
  if (clearTimeWindowBtn) {{
      clearTimeWindowBtn.addEventListener('click', (e) => {{
          e.preventDefault();
          clearTimeWindow();
      }});
  }}

}}); // DOMContentLoaded
</script>

</body>
</html>
"""
 # Per-transaction rows (Transaction Statistics table) - Use ALL labels
    per_label_rows_html = []
    for ls in per_label:
        avg = ls.get("avg")
        err_rate = ls.get("error_rate")
        count = ls.get("count", 0)
        tps = (count / duration_sec_val) if duration_sec_val > 0 else 0.0
        throughput_per_min = tps * 60.0
        err_pct = (err_rate * 100.0) if err_rate is not None else 0.0

        per_label_rows_html.append(f"""
        <tr data-avg-ms="{avg if avg is not None else ''}"
            data-error-pct="{err_pct if err_rate is not None else ''}"
            data-tps="{tps}">
          <td>{ls["label"]}</td>
          <td>{ls["count"]}</td>
          <td>{format_ms(ls["min"])}</td>
          <td>{format_ms(ls["avg"])}</td>
          <td>{format_ms(ls["p50"])}</td>
          <td>{format_ms(ls["p90"])}</td>
          <td>{format_ms(ls["p95"])}</td>
          <td>{format_ms(ls["p99"])}</td>
          <td>{format_ms(ls["max"])}</td>
          <td>{throughput_per_min:.2f}</td>
          <td>{tps:.3f}</td>
          <td>{format_pct(ls["error_rate"])}</td>
        </tr>
        """)
    per_label_rows_html_str = "\n".join(per_label_rows_html)

    # Error rows (summary by response code)
    error_rows_html = []
    if error_summary:
        for er in error_summary:
            error_rows_html.append(f"""
            <tr>
              <td>{er["response_code"]}</td>
              <td>{er["count"]}</td>
              <td>{format_pct(er["pct"])}</td>
            </tr>
            """)
    else:
        error_rows_html.append(
            '<tr><td colspan="3" class="small">No errors recorded.</td></tr>'
        )
    error_rows_html_str = "\n".join(error_rows_html)

    # Error details rows (transaction + message)
    error_detail_rows_html = []
    if error_details:
        for ed in error_details:
            msg = ed["response_message"] or ""
            msg = msg.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            error_detail_rows_html.append(f"""
            <tr>
              <td>{ed["label"]}</td>
              <td>{ed["response_code"]}</td>
              <td>{msg}</td>
              <td>{ed["count"]}</td>
              <td>{format_pct(ed["pct"])}</td>
            </tr>
            """)
    else:
        error_detail_rows_html.append(
            '<tr><td colspan="5" class="small">No detailed error records available.</td></tr>'
        )
    error_detail_rows_html_str = "\n".join(error_detail_rows_html)

    sla_total = sla_result["total_checks"]
    sla_passed = sla_result["passed_checks"]
    sla_pct = sla_result["overall_pct"] * 100.0
    sla_overall_class = "sla-overall-ok" if sla_result["overall_status"] else "sla-overall-fail"
    sla_overall_text = "PASS" if sla_result["overall_status"] else "FAIL"

    html = html_template.format(
        test_name=test_name,
        environment=environment,
        generated_at=generated_at,
        start_time=start_dt.strftime("%Y-%m-%d %H:%M:%S"),
        end_time=end_dt.strftime("%Y-%m-%d %H:%M:%S"),
        duration_str=duration_str,
        total_requests=global_stats["total_requests"],
        duration_sec=global_stats["duration_sec"],
        peak_rps=metrics["peak_rps"],
        avg_ms=global_stats["avg_ms"],
        p95_ms=global_stats["p95_ms"],
        error_rate_pct=global_stats["error_rate"] * 100.0,
        apdex_value=global_stats["apdex"],
        apdex_threshold=apdex_threshold_s,
        successes=global_stats["successes"],
        errors=global_stats["errors"],
        error_rows_html=error_rows_html_str,
        error_detail_rows_html=error_detail_rows_html_str,
        per_label_rows_html=per_label_rows_html_str,
        transaction_sla_rows_html=transaction_sla_rows_html_str,
        chart_data_json=chart_data_json,
        per_label_json=per_label_json,
        baseline_json=baseline_json,
        infra_data_json=infra_data_json,
        sla_total=sla_total,
        sla_passed=sla_passed,
        sla_pct=sla_pct,
        sla_overall_class=sla_overall_class,
        sla_overall_text=sla_overall_text,
        overall_insights_html=overall_insights_html,
        infra_insights_html=infra_insights_html,
    )
    return html


# =========================
# SUMMARY JSON GENERATION
# =========================

def generate_summary_json(metrics: Dict[str, Any],
                          sla_result: Dict[str, Any],
                          health_score: Dict[str, Any],
                          output_path: str) -> None:
    """Generate a JSON summary of the test results."""
    summary = {
        "metadata": {
            "generated_at": datetime.now().isoformat(),
            "version": "3.8"
        },
        "test_summary": {
            "total_requests": metrics["global"]["total_requests"],
            "duration_sec": metrics["global"]["duration_sec"],
            "error_rate": metrics["global"]["error_rate"],
            "avg_response_ms": metrics["global"]["avg_ms"],
            "p95_response_ms": metrics["global"]["p95_ms"],
            "peak_rps": metrics["peak_rps"]
        },
        "sla_results": {
            "passed_checks": sla_result["passed_checks"],
            "total_checks": sla_result["total_checks"],
            "pass_rate": sla_result["overall_pct"],
            "status": "PASS" if sla_result["overall_status"] else "FAIL"
        },
        "health_score": health_score
    }

    # Create directory if it doesn't exist
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(summary, f, indent=2, default=str)

    print(f"✓ Summary JSON generated: {output_path}")


# =========================
# CLI
# =========================

def main():
    parser = argparse.ArgumentParser(
        description="Generate an enhanced HTML Performance Dashboard from JMeter CSV/JTL with advanced insights.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Simple with config file
  python jmeter-report.py --config config.yaml

  # Custom with config file and command line overrides
  python jmeter-report.py --config config.yaml --output custom_report.html

  # Traditional command line (backwards compatible)
  python jmeter-report.py --input results.csv --output report.html --test-name "My Test"

  # Config file only
  python jmeter-report.py --config quick_test.yaml
        """
    )

    # Configuration file option (NEW)
    parser.add_argument("--config",
                        help="Path to configuration file (JSON or YAML). "
                             "All other arguments can be specified in this file.")

    # Keep all existing arguments but make most optional when config is provided
    parser.add_argument("--input", help="Path to JMeter CSV/JTL results file")
    parser.add_argument("--output", help="Output HTML report path")
    parser.add_argument("--test-name", default="JMeter Load Test",
                        help="Name of the test to display in the report")
    parser.add_argument("--environment", default="Unknown",
                        help="Environment (e.g. QA, UAT, Prod-like)")
    parser.add_argument("--apdex-threshold", type=float, default=1.5,
                        help="APDEX threshold T in seconds (default 1.5)")
    parser.add_argument("--baseline",
                        help="Optional baseline JMeter CSV/JTL for comparison")
    parser.add_argument("--sla-config-file",
                        help="Path to SLA config JSON file")
    parser.add_argument("--sla-config-json",
                        help="SLA config as JSON string")
    parser.add_argument("--infra-json", nargs="*",
                        help="Optional JSON files with infrastructure metrics")
    parser.add_argument("--excel-export",
                        help="Export to Excel file (provide path)")
    parser.add_argument("--excel-only",
                        action="store_true",
                        help="Generate only Excel report, no HTML")
    parser.add_argument("--enable-container-analysis",
                        action="store_true",
                        default=True,
                        help="Enable container resource analysis (default: True)")

    args = parser.parse_args()

    # Load config file if provided
    config = {}
    if args.config:
        if not os.path.exists(args.config):
            raise SystemExit(f"Error: Config file not found: {args.config}")
        config = load_config_file(args.config)
        # Merge config with command line args (command line takes precedence)
        args = merge_config_with_args(config, args)

    # Validate required arguments
    if not args.input:
        if args.config:
            raise SystemExit(f"Error: Input file not specified in config file or command line")
        else:
            parser.error("--input is required when not using --config")

    if not args.excel_only and not args.output:
        if args.config and 'output' in config.get('output', {}):
            # Output specified in config
            pass
        else:
            parser.error("--output is required unless --excel-only is specified")

    if args.excel_only and not args.excel_export:
        parser.error("--excel-export is required when using --excel-only")

    # Apply timestamp substitution in output paths
    if args.output and '${TIMESTAMP}' in args.output:
        timestamp_format = config.get('output', {}).get('timestamp_format', '%Y%m%d_%H%M%S')
        timestamp = datetime.now().strftime(timestamp_format)
        args.output = args.output.replace('${TIMESTAMP}', timestamp)

    if args.excel_export and '${TIMESTAMP}' in args.excel_export:
        timestamp_format = config.get('output', {}).get('timestamp_format', '%Y%m%d_%H%M%S')
        timestamp = datetime.now().strftime(timestamp_format)
        args.excel_export = args.excel_export.replace('${TIMESTAMP}', timestamp)

    # Print configuration summary
    print(f"\n{'=' * 60}")
    print(f"JMeter Performance Report v3.8")
    print(f"{'=' * 60}")
    if args.config:
        print(f"Configuration file: {args.config}")
    print(f"Input file:        {args.input}")
    if args.baseline:
        print(f"Baseline file:     {args.baseline}")
    if not args.excel_only:
        print(f"HTML output:       {args.output}")
    if args.excel_export:
        print(f"Excel output:      {args.excel_export}")
    print(f"Test name:         {args.test_name}")
    print(f"Environment:       {args.environment}")
    print(f"{'=' * 60}\n")

    # SLA configuration
    sla_config = DEFAULT_SLA_CONFIG
    if args.sla_config_file:
        try:
            with open(args.sla_config_file, "r", encoding="utf-8") as f:
                sla_config = json.load(f)
        except Exception as e:
            raise SystemExit(f"Failed to load SLA config from file: {e}")
    elif args.sla_config_json:
        try:
            sla_config = json.loads(args.sla_config_json)
        except Exception as e:
            raise SystemExit(f"Failed to parse SLA config JSON: {e}")

    # Parse and compute metrics
    samples = parse_jmeter_csv(args.input)
    if not samples:
        raise SystemExit("No samples parsed from input file. Check CSV/JTL format.")

    metrics = compute_metrics(samples, args.apdex_threshold)

    # Extract transaction controller names
    try:
        metrics["transaction_names"] = extract_transaction_names(samples)
    except Exception:
        metrics["transaction_names"] = []

    sla_result = evaluate_sla(metrics, sla_config)

    baseline_comparison = None
    if args.baseline:
        baseline_samples = parse_jmeter_csv(args.baseline)
        if baseline_samples:
            baseline_metrics = compute_metrics(baseline_samples, args.apdex_threshold)
            baseline_comparison = compare_baseline(metrics, baseline_metrics)

    infra_metrics = None
    if getattr(args, "infra_json", None):
        infra_metrics = build_infra_metrics(args.infra_json)

    # Container comparison analysis
    container_comparison = None
    if infra_metrics and args.enable_container_analysis:
        container_comparison = analyze_container_comparison(infra_metrics)

    # Calculate health score for summary
    health_score = calculate_health_score(metrics['global'], sla_result, metrics)

    # Generate summary JSON if requested in config
    if config.get('options', {}).get('generate_summary_json', False):
        summary_path = config['options'].get('summary_json', 'report_summary.json')
        generate_summary_json(metrics, sla_result, health_score, summary_path)

    # Excel Export
    if args.excel_export:
        if not EXCEL_SUPPORT:
            print("Error: Excel export requires openpyxl. Install with: pip install openpyxl")
            if args.excel_only:
                raise SystemExit("Cannot generate Excel report without openpyxl")
        else:
            try:
                excel_path = export_to_excel(
                    metrics=metrics,
                    sla_result=sla_result,
                    container_comparison=container_comparison,
                    test_name=args.test_name,
                    environment=args.environment,
                    output_path=args.excel_export
                )
                print(f"✓ Excel report generated: {excel_path}")
            except Exception as e:
                print(f"✗ Excel export failed: {e}")
                if args.excel_only:
                    raise SystemExit("Excel export failed")

    # HTML generation (unless excel-only)
    if not args.excel_only:
        html = build_html(
            metrics=metrics,
            sla_result=sla_result,
            sla_config=sla_config,
            test_name=args.test_name,
            environment=args.environment,
            apdex_threshold_s=args.apdex_threshold,
            baseline_comparison=baseline_comparison,
            infra_metrics=infra_metrics,
            container_comparison=container_comparison,
        )

        # Create output directory if it doesn't exist
        os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)

        with open(args.output, "w", encoding="utf-8") as f:
            f.write(html)

        print(f"\n✓ Enhanced HTML report v3.8 generated: {args.output}")
        print(f"  - Health Score calculation with letter grade")
        print(f"  - Root Cause Analysis with confidence levels")
        if container_comparison:
            print(f"  - Container comparison analysis")
        print(f"  - Priority-based recommendations")
        print(f"  - Capacity planning estimates")

        if infra_metrics:
            print(f"  - Infrastructure metrics integrated")
        if args.excel_export:
            print(f"  - Excel report also available: {args.excel_export}")

        if config.get('options', {}).get('generate_summary_json', False):
            print(f"  - Summary JSON generated")

if __name__ == "__main__":
    main()